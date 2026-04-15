#!/usr/bin/env python3
"""
4_bilan_carbone.py — Calculer le bilan carbone (IC construction) d'une maquette IFC
Usage : python 4_bilan_carbone.py chemin/vers/maquette.ifc

Base de donnees FDES simplifiee (valeurs indicatives issues INIES).
Pour un bilan officiel, utiliser les vraies FDES INIES.

Produit : fichier Excel avec :
  - IC par lot (kg CO2 eq)
  - IC par m² SHON
  - Comparaison avec seuils RE2020
  - Preconisations
"""
import sys
import os


# Base de donnees simplifiee - valeurs INIES indicatives en kg CO2 eq par unite
# IMPORTANT : ces valeurs sont indicatives. Utiliser les vraies FDES pour un projet reel.
FDES_SIMPLE = {
    # Materiaux / Unite / kg CO2 eq par unite
    'beton_standard_c25_m3': 250,           # m³
    'beton_bas_carbone_m3': 180,            # m³
    'acier_ba_kg': 2.0,                     # kg (≈ 150 kg/m³ de beton arme)
    'parpaing_m2': 15,                      # m²
    'brique_creuse_m2': 12,                 # m²
    'brique_monomur_m2': 25,                # m² (inclut isolation)
    'placo_ba13_m2': 3,                     # m²
    'laine_verre_140mm_m2': 6,              # m²
    'laine_roche_140mm_m2': 10,             # m²
    'fibre_bois_140mm_m2': -8,              # m² (negatif = stockage carbone)
    'pse_100mm_m2': 12,                     # m²
    'ouate_cellulose_m2': -5,               # m² (stockage)
    'charpente_bois_m2': -25,               # m² (stockage)
    'charpente_metal_m2': 40,               # m²
    'couverture_tuiles_m2': 18,             # m²
    'couverture_bac_acier_m2': 15,          # m²
    'menuiserie_pvc_m2': 90,                # m² de fenetre
    'menuiserie_alu_m2': 130,               # m²
    'menuiserie_bois_m2': -10,              # m² (stockage)
    'porte_bois_u': 50,                     # unite
    'porte_acier_u': 120,                   # unite
}


def calculer_bilan_carbone(chemin_ifc: str):
    try:
        import ifcopenshell
        import ifcopenshell.util.element
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"[ERREUR] Bibliotheque manquante : {e}")
        sys.exit(1)

    if not os.path.exists(chemin_ifc):
        print(f"[ERREUR] Fichier introuvable : {chemin_ifc}")
        sys.exit(1)

    print(f"\nBilan carbone : {os.path.basename(chemin_ifc)}")
    model = ifcopenshell.open(chemin_ifc)

    def get_qto(element, prop_name):
        psets = ifcopenshell.util.element.get_psets(element, qtos_only=True)
        for qto_name, props in psets.items():
            if prop_name in props:
                return props[prop_name]
        return 0

    # Extraire les quantites
    vol_fond = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                   for e in (model.by_type("IfcFooting") + model.by_type("IfcPile")))
    vol_mur_ba = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                     for e in model.by_type("IfcWall"))
    surf_mur = sum((get_qto(e, 'NetSideArea') or get_qto(e, 'GrossSideArea') or 0)
                   for e in model.by_type("IfcWall"))
    vol_dalle = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                    for e in model.by_type("IfcSlab"))
    surf_dalle = sum((get_qto(e, 'NetArea') or get_qto(e, 'GrossArea') or 0)
                     for e in model.by_type("IfcSlab"))
    vol_pot = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                  for e in model.by_type("IfcColumn"))
    vol_pou = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                  for e in model.by_type("IfcBeam"))
    surf_toit = sum((get_qto(e, 'NetArea') or get_qto(e, 'GrossArea') or 0)
                    for e in model.by_type("IfcRoof"))
    nb_portes = len(model.by_type("IfcDoor"))
    nb_fenetres = len(model.by_type("IfcWindow"))

    # SHON approx = surface des dalles (hors toiture)
    shon = max(surf_dalle - surf_toit, surf_dalle / 2) if surf_dalle > 0 else 100

    # Volume beton total pour l'acier
    vol_beton_total = vol_fond + vol_mur_ba + vol_dalle + vol_pot + vol_pou
    # 150 kg d'acier par m³ de BA (ratio moyen)
    masse_acier = vol_beton_total * 150

    # Calcul du carbone par lot
    lots = {
        'LOT GROS OEUVRE': [
            ('Beton fondations', vol_fond, 'm³', FDES_SIMPLE['beton_standard_c25_m3']),
            ('Beton murs', vol_mur_ba, 'm³', FDES_SIMPLE['beton_standard_c25_m3']),
            ('Beton dalles', vol_dalle, 'm³', FDES_SIMPLE['beton_standard_c25_m3']),
            ('Beton poteaux', vol_pot, 'm³', FDES_SIMPLE['beton_standard_c25_m3']),
            ('Beton poutres', vol_pou, 'm³', FDES_SIMPLE['beton_standard_c25_m3']),
            ('Acier beton arme', masse_acier, 'kg', FDES_SIMPLE['acier_ba_kg']),
            ('Maconnerie (estimation)', surf_mur * 0.3, 'm²', FDES_SIMPLE['parpaing_m2']),
        ],
        'LOT CHARPENTE / COUVERTURE': [
            ('Charpente (bois estime)', surf_toit, 'm²', FDES_SIMPLE['charpente_bois_m2']),
            ('Couverture (tuiles)', surf_toit, 'm²', FDES_SIMPLE['couverture_tuiles_m2']),
        ],
        'LOT MENUISERIES': [
            ('Fenetres PVC (estimation 1.5 m² moyenne)', nb_fenetres * 1.5, 'm²', FDES_SIMPLE['menuiserie_pvc_m2']),
            ('Portes bois', nb_portes * 0.7, 'U', FDES_SIMPLE['porte_bois_u']),
        ],
        'LOT CLOISONS / DOUBLAGES': [
            ('Cloisons placo BA13', surf_mur * 0.4, 'm²', FDES_SIMPLE['placo_ba13_m2']),
            ('Doublages isolants laine verre', surf_mur * 0.3, 'm²', FDES_SIMPLE['laine_verre_140mm_m2']),
        ],
    }

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilan Carbone"

    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="166534")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="166534")
    lot_font = Font(bold=True, size=12, color="FFFFFF")
    lot_fill = PatternFill("solid", fgColor="475569")

    # En-tete
    ws.merge_cells('A1:F1')
    ws['A1'] = "BILAN CARBONE - IC Construction"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"Projet : {os.path.splitext(os.path.basename(chemin_ifc))[0]}"
    ws['A3'] = f"SHON estimee : {shon:.0f} m²"
    ws['A4'] = "Source : FDES simplifiees (valeurs indicatives - INIES)"
    ws['A4'].font = Font(italic=True, color="666666")

    # Headers
    row = 6
    headers = ["Ouvrage", "Quantite", "Unite", "Facteur (kg CO2eq/U)", "Carbone (kg CO2eq)", "Note"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    total_general = 0
    totaux_par_lot = {}

    for nom_lot, items in lots.items():
        # Ligne du lot
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=nom_lot)
        cell.font = lot_font
        cell.fill = lot_fill
        cell.alignment = Alignment(horizontal='left', indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        total_lot = 0
        for ouvrage, qte, unite, facteur in items:
            carbone = qte * facteur
            total_lot += carbone
            note = "✓ Stockage" if facteur < 0 else ""

            ws.cell(row=row, column=1, value=ouvrage)
            ws.cell(row=row, column=2, value=round(qte, 2))
            ws.cell(row=row, column=3, value=unite)
            ws.cell(row=row, column=4, value=facteur)
            ws.cell(row=row, column=5, value=round(carbone, 0))
            ws.cell(row=row, column=6, value=note)

            if facteur < 0:
                ws.cell(row=row, column=5).font = Font(color="16A34A")
            row += 1

        # Sous-total du lot
        ws.cell(row=row, column=1, value=f"Sous-total {nom_lot}").font = Font(bold=True)
        ws.cell(row=row, column=5, value=round(total_lot, 0)).font = Font(bold=True)
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="E2E8F0")
        totaux_par_lot[nom_lot] = total_lot
        total_general += total_lot
        row += 2

    # Total general
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="TOTAL IC Construction (kg CO2 eq)")
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="991B1B")
    cell.alignment = Alignment(horizontal='right', indent=1)
    total_cell = ws.cell(row=row, column=5, value=round(total_general, 0))
    total_cell.font = Font(bold=True, size=14)
    total_cell.fill = PatternFill("solid", fgColor="FEF3C7")
    ws.row_dimensions[row].height = 28
    row += 1

    # Ratio par m² SHON
    ic_par_m2 = total_general / shon if shon > 0 else 0
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="IC / m² SHON").font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(ic_par_m2, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value="kg CO2 eq/m²")
    row += 2

    # Comparaison RE2020
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="COMPARAISON SEUILS RE2020 (logement individuel)")
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal='center')
    row += 1

    seuils = [
        ("Seuil RE2020 2022-2024", 640, "Respecte" if ic_par_m2 <= 640 else "Depasse"),
        ("Seuil RE2020 2025-2027", 530, "Respecte" if ic_par_m2 <= 530 else "Depasse"),
        ("Seuil RE2020 2028-2030", 475, "Respecte" if ic_par_m2 <= 475 else "Depasse"),
        ("Seuil RE2020 2031+", 415, "Respecte" if ic_par_m2 <= 415 else "Depasse"),
    ]

    for nom, seuil, statut in seuils:
        ws.cell(row=row, column=1, value=nom)
        ws.cell(row=row, column=2, value=seuil)
        ws.cell(row=row, column=3, value="kg CO2 eq/m²")
        ws.cell(row=row, column=4, value=round(ic_par_m2, 2))
        ws.cell(row=row, column=5, value=statut)
        if statut == "Respecte":
            ws.cell(row=row, column=5).font = Font(bold=True, color="16A34A")
        else:
            ws.cell(row=row, column=5).font = Font(bold=True, color="DC2626")
        row += 1

    # Preconisations
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="PRECONISATIONS POUR REDUIRE L'IMPACT CARBONE")
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="059669")
    cell.alignment = Alignment(horizontal='center')
    row += 1

    preconisations = [
        "1. Remplacer le beton standard par du beton bas carbone (-28% CO2)",
        "2. Utiliser des isolants biosources (fibre de bois, ouate de cellulose) : stockage carbone",
        "3. Privilegier la charpente bois (stockage carbone) plutot que metal",
        "4. Menuiseries bois ou alu recycle (eviter alu vierge)",
        "5. Maconnerie en brique monomur (integre l'isolation, moins de materiaux)",
        "6. Reduire le volume de beton (dalles optimisees, fondations peu profondes)",
        "7. Integrer de la preparation hors site / prefabrication (moins de dechets)",
    ]

    for p in preconisations:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, indent=1)
        row += 1

    # Largeurs colonnes
    widths = [45, 15, 10, 18, 20, 25]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sauvegarder
    base = os.path.splitext(chemin_ifc)[0]
    output = base + "_bilan_carbone.xlsx"
    wb.save(output)

    print(f"\n[OK] Bilan carbone genere : {output}")
    print(f"\nResultats cles :")
    print(f"  Total IC construction : {total_general:.0f} kg CO2 eq")
    print(f"  Par m² SHON           : {ic_par_m2:.1f} kg CO2 eq/m²")
    print(f"  Seuil RE2020 2025     : 530 kg CO2 eq/m²")
    if ic_par_m2 <= 530:
        print(f"  [OK] Projet conforme aux seuils RE2020 2025")
    else:
        print(f"  [!]  Projet au-dessus du seuil RE2020 2025 (voir preconisations)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python 4_bilan_carbone.py <chemin_vers_maquette.ifc>")
        sys.exit(1)
    calculer_bilan_carbone(sys.argv[1])
