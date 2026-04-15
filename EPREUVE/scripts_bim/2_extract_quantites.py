#!/usr/bin/env python3
"""
2_extract_quantites.py — Extraire les quantites d'une maquette IFC vers Excel
Usage : python 2_extract_quantites.py chemin/vers/maquette.ifc

Produit : fichier Excel avec plusieurs feuilles :
  - Murs (surface, volume, longueur, porteur/non-porteur, ext/int)
  - Dalles (surface, volume, epaisseur, etage)
  - Poteaux (volume, hauteur, etage)
  - Poutres (volume, longueur, etage)
  - Fondations (type, volume)
  - Portes et fenetres (dimensions, comptage)
  - Synthese (totaux par lot)
"""
import sys
import os
from collections import defaultdict


def extraire_quantites(chemin_ifc: str):
    try:
        import ifcopenshell
        import ifcopenshell.util.element
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError as e:
        print(f"[ERREUR] Bibliotheque manquante : {e}")
        print("Lancez : pip install -r requirements.txt")
        sys.exit(1)

    if not os.path.exists(chemin_ifc):
        print(f"[ERREUR] Fichier introuvable : {chemin_ifc}")
        sys.exit(1)

    print(f"\nExtraction des quantites : {os.path.basename(chemin_ifc)}")
    model = ifcopenshell.open(chemin_ifc)

    # Workbook Excel
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

    def get_qto(element, prop_name):
        """Extrait une quantite depuis les QuantitySets IFC."""
        psets = ifcopenshell.util.element.get_psets(element, qtos_only=True)
        for qto_name, props in psets.items():
            if prop_name in props:
                return props[prop_name]
        return None

    def get_pset_value(element, pset_name, prop_name):
        psets = ifcopenshell.util.element.get_psets(element)
        return psets.get(pset_name, {}).get(prop_name)

    def get_etage(element):
        """Trouve l'etage d'un element."""
        storey = ifcopenshell.util.element.get_container(element)
        if storey and storey.is_a('IfcBuildingStorey'):
            return storey.Name or 'N/C'
        return 'N/C'

    # =========================================================
    # FEUILLE 1 : MURS
    # =========================================================
    ws = wb.create_sheet("Murs")
    style_header(ws, ["Nom", "Type", "Etage", "Longueur (m)", "Hauteur (m)",
                      "Surface (m²)", "Volume (m³)", "Porteur", "Exterieur", "Materiau"])

    murs = model.by_type("IfcWall")
    row = 2
    total_vol_mur = 0
    total_surf_mur = 0
    for m in murs:
        longueur = get_qto(m, 'Length') or 0
        hauteur = get_qto(m, 'Height') or 0
        surface = get_qto(m, 'NetSideArea') or get_qto(m, 'GrossSideArea') or 0
        volume = get_qto(m, 'NetVolume') or get_qto(m, 'GrossVolume') or 0
        porteur = get_pset_value(m, 'Pset_WallCommon', 'LoadBearing')
        exterieur = get_pset_value(m, 'Pset_WallCommon', 'IsExternal')

        # Materiau
        materiau = ''
        try:
            for rel in model.by_type('IfcRelAssociatesMaterial'):
                if m in rel.RelatedObjects:
                    mat = rel.RelatingMaterial
                    if hasattr(mat, 'Name'):
                        materiau = mat.Name or ''
                        break
        except:
            pass

        ws.cell(row=row, column=1, value=m.Name or f"Mur_{m.id()}")
        ws.cell(row=row, column=2, value=m.is_a())
        ws.cell(row=row, column=3, value=get_etage(m))
        ws.cell(row=row, column=4, value=round(longueur, 3))
        ws.cell(row=row, column=5, value=round(hauteur, 3))
        ws.cell(row=row, column=6, value=round(surface, 3))
        ws.cell(row=row, column=7, value=round(volume, 3))
        ws.cell(row=row, column=8, value="Oui" if porteur else "Non")
        ws.cell(row=row, column=9, value="Oui" if exterieur else "Non")
        ws.cell(row=row, column=10, value=materiau)
        total_vol_mur += volume
        total_surf_mur += surface
        row += 1

    # Ligne de totaux
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_surf_mur, 2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_vol_mur, 2)).font = Font(bold=True)
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 15

    print(f"  Murs         : {len(murs)} ({total_vol_mur:.2f} m³, {total_surf_mur:.2f} m²)")

    # =========================================================
    # FEUILLE 2 : DALLES
    # =========================================================
    ws = wb.create_sheet("Dalles")
    style_header(ws, ["Nom", "Type", "Etage", "Surface (m²)", "Volume (m³)",
                      "Epaisseur (m)", "Porteur", "Materiau"])

    dalles = model.by_type("IfcSlab")
    row = 2
    total_vol_dalle = 0
    total_surf_dalle = 0
    for d in dalles:
        surface = get_qto(d, 'NetArea') or get_qto(d, 'GrossArea') or 0
        volume = get_qto(d, 'NetVolume') or get_qto(d, 'GrossVolume') or 0
        epaisseur = get_qto(d, 'Depth') or 0
        porteur = get_pset_value(d, 'Pset_SlabCommon', 'LoadBearing')
        pred = get_pset_value(d, 'Pset_SlabCommon', 'PredefinedType') or ''

        ws.cell(row=row, column=1, value=d.Name or f"Dalle_{d.id()}")
        ws.cell(row=row, column=2, value=pred or d.is_a())
        ws.cell(row=row, column=3, value=get_etage(d))
        ws.cell(row=row, column=4, value=round(surface, 3))
        ws.cell(row=row, column=5, value=round(volume, 3))
        ws.cell(row=row, column=6, value=round(epaisseur, 3))
        ws.cell(row=row, column=7, value="Oui" if porteur else "Non")
        total_vol_dalle += volume
        total_surf_dalle += surface
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_surf_dalle, 2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_vol_dalle, 2)).font = Font(bold=True)
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15

    print(f"  Dalles       : {len(dalles)} ({total_vol_dalle:.2f} m³, {total_surf_dalle:.2f} m²)")

    # =========================================================
    # FEUILLE 3 : POTEAUX
    # =========================================================
    ws = wb.create_sheet("Poteaux")
    style_header(ws, ["Nom", "Etage", "Hauteur (m)", "Volume (m³)", "Section"])

    poteaux = model.by_type("IfcColumn")
    row = 2
    total_vol_pot = 0
    for p in poteaux:
        hauteur = get_qto(p, 'Length') or 0
        volume = get_qto(p, 'NetVolume') or get_qto(p, 'GrossVolume') or 0
        section = get_qto(p, 'CrossSectionArea') or 0

        ws.cell(row=row, column=1, value=p.Name or f"Poteau_{p.id()}")
        ws.cell(row=row, column=2, value=get_etage(p))
        ws.cell(row=row, column=3, value=round(hauteur, 3))
        ws.cell(row=row, column=4, value=round(volume, 3))
        ws.cell(row=row, column=5, value=round(section, 4))
        total_vol_pot += volume
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_vol_pot, 2)).font = Font(bold=True)
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 15

    print(f"  Poteaux      : {len(poteaux)} ({total_vol_pot:.2f} m³)")

    # =========================================================
    # FEUILLE 4 : POUTRES
    # =========================================================
    ws = wb.create_sheet("Poutres")
    style_header(ws, ["Nom", "Etage", "Longueur (m)", "Volume (m³)", "Section"])

    poutres = model.by_type("IfcBeam")
    row = 2
    total_vol_pou = 0
    for p in poutres:
        longueur = get_qto(p, 'Length') or 0
        volume = get_qto(p, 'NetVolume') or get_qto(p, 'GrossVolume') or 0
        section = get_qto(p, 'CrossSectionArea') or 0

        ws.cell(row=row, column=1, value=p.Name or f"Poutre_{p.id()}")
        ws.cell(row=row, column=2, value=get_etage(p))
        ws.cell(row=row, column=3, value=round(longueur, 3))
        ws.cell(row=row, column=4, value=round(volume, 3))
        ws.cell(row=row, column=5, value=round(section, 4))
        total_vol_pou += volume
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_vol_pou, 2)).font = Font(bold=True)
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 15

    print(f"  Poutres      : {len(poutres)} ({total_vol_pou:.2f} m³)")

    # =========================================================
    # FEUILLE 5 : FONDATIONS
    # =========================================================
    ws = wb.create_sheet("Fondations")
    style_header(ws, ["Nom", "Type", "Volume (m³)"])

    fondations = model.by_type("IfcFooting") + model.by_type("IfcPile")
    row = 2
    total_vol_fond = 0
    for f in fondations:
        volume = get_qto(f, 'NetVolume') or get_qto(f, 'GrossVolume') or 0

        ws.cell(row=row, column=1, value=f.Name or f"Fondation_{f.id()}")
        ws.cell(row=row, column=2, value=f.is_a())
        ws.cell(row=row, column=3, value=round(volume, 3))
        total_vol_fond += volume
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_vol_fond, 2)).font = Font(bold=True)
    for col in range(1, 4):
        ws.column_dimensions[chr(64 + col)].width = 18

    print(f"  Fondations   : {len(fondations)} ({total_vol_fond:.2f} m³)")

    # =========================================================
    # FEUILLE 6 : PORTES / FENETRES
    # =========================================================
    ws = wb.create_sheet("Menuiseries")
    style_header(ws, ["Type", "Nom", "Etage", "Largeur (m)", "Hauteur (m)", "Surface (m²)"])

    row = 2
    nb_portes = 0
    nb_fenetres = 0

    for p in model.by_type("IfcDoor"):
        ws.cell(row=row, column=1, value="Porte")
        ws.cell(row=row, column=2, value=p.Name or f"Porte_{p.id()}")
        ws.cell(row=row, column=3, value=get_etage(p))
        largeur = getattr(p, 'OverallWidth', 0) or 0
        hauteur = getattr(p, 'OverallHeight', 0) or 0
        ws.cell(row=row, column=4, value=round(largeur, 3))
        ws.cell(row=row, column=5, value=round(hauteur, 3))
        ws.cell(row=row, column=6, value=round(largeur * hauteur, 3))
        nb_portes += 1
        row += 1

    for f in model.by_type("IfcWindow"):
        ws.cell(row=row, column=1, value="Fenetre")
        ws.cell(row=row, column=2, value=f.Name or f"Fenetre_{f.id()}")
        ws.cell(row=row, column=3, value=get_etage(f))
        largeur = getattr(f, 'OverallWidth', 0) or 0
        hauteur = getattr(f, 'OverallHeight', 0) or 0
        ws.cell(row=row, column=4, value=round(largeur, 3))
        ws.cell(row=row, column=5, value=round(hauteur, 3))
        ws.cell(row=row, column=6, value=round(largeur * hauteur, 3))
        nb_fenetres += 1
        row += 1

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 15

    print(f"  Portes       : {nb_portes}")
    print(f"  Fenetres     : {nb_fenetres}")

    # =========================================================
    # FEUILLE 7 : SYNTHESE
    # =========================================================
    ws = wb.create_sheet("Synthese", 0)  # En premiere position
    ws['A1'] = "SYNTHESE QUANTITATIVE - Lot Gros Oeuvre"
    ws['A1'].font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells('A1:D1')

    headers = ["Ouvrage", "Unite", "Quantite", "Observation"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    data = [
        ("Fondations (toutes)", "m³", round(total_vol_fond, 2), "Semelles + pieux"),
        ("Murs (volume beton)", "m³", round(total_vol_mur, 2), f"{len(murs)} murs"),
        ("Murs (surface)", "m²", round(total_surf_mur, 2), ""),
        ("Dalles (volume beton)", "m³", round(total_vol_dalle, 2), f"{len(dalles)} dalles"),
        ("Dalles (surface)", "m²", round(total_surf_dalle, 2), ""),
        ("Poteaux (volume beton)", "m³", round(total_vol_pot, 2), f"{len(poteaux)} poteaux"),
        ("Poutres (volume beton)", "m³", round(total_vol_pou, 2), f"{len(poutres)} poutres"),
        ("", "", "", ""),
        ("TOTAL BETON ARME", "m³",
         round(total_vol_fond + total_vol_mur + total_vol_dalle + total_vol_pot + total_vol_pou, 2),
         "Somme (verification visuelle requise)"),
        ("", "", "", ""),
        ("Portes", "U", nb_portes, ""),
        ("Fenetres", "U", nb_fenetres, ""),
    ]

    for i, (ouvrage, unite, qte, obs) in enumerate(data, 4):
        ws.cell(row=i, column=1, value=ouvrage)
        ws.cell(row=i, column=2, value=unite)
        ws.cell(row=i, column=3, value=qte)
        ws.cell(row=i, column=4, value=obs)
        if "TOTAL" in str(ouvrage):
            for col in range(1, 5):
                ws.cell(row=i, column=col).font = Font(bold=True)
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="FEF3C7")

    for col, width in enumerate([30, 10, 15, 35], 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # Sauvegarder
    base = os.path.splitext(chemin_ifc)[0]
    output = base + "_quantites.xlsx"
    wb.save(output)

    print(f"\n[OK] Fichier genere : {output}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python 2_extract_quantites.py <chemin_vers_maquette.ifc>")
        sys.exit(1)
    extraire_quantites(sys.argv[1])
