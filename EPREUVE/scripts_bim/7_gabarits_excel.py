#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
7_gabarits_excel.py - Genere des gabarits Excel modernes pour l'examen E6-A BIM
Usage : python 7_gabarits_excel.py [dossier_sortie]

Genere 5 gabarits pret-a-remplir :
  1. Dashboard projet (KPI globaux)
  2. DPGF professionnel (12 lots + totaux auto)
  3. Planning chantier (Gantt simplifie)
  4. Bilan carbone RE2020 (jauges + seuils)
  5. Controle qualite maquette BIM (checklist + score)
"""
import sys
import os

from _utils import setup_encoding

setup_encoding()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
    from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("[ERREUR] openpyxl manquant. Installez avec : python -m pip install openpyxl")
    sys.exit(1)


# ============================================================
# PALETTE COULEURS MODERNE
# ============================================================
NAVY = "1E3A5F"
NAVY_DARK = "0F1E33"
ORANGE = "F58220"
TURQUOISE = "4DC7C7"
GREEN = "16A34A"
RED = "DC2626"
YELLOW = "F59E0B"
GREY_LIGHT = "F1F5F9"
GREY = "94A3B8"
WHITE = "FFFFFF"


# ============================================================
# HELPERS STYLES
# ============================================================
def thin_border():
    s = Side(style='thin', color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def style_titre(ws, cell_range, texte, bg=NAVY, fg=WHITE, taille=16, hauteur=32):
    ws.merge_cells(cell_range)
    first = cell_range.split(':')[0]
    c = ws[first]
    c.value = texte
    c.font = Font(bold=True, size=taille, color=fg, name="Calibri")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[c.row].height = hauteur


def style_sous_titre(ws, row, cell_range, texte, bg=ORANGE, fg=WHITE):
    ws.merge_cells(cell_range)
    first = cell_range.split(':')[0]
    c = ws[first]
    c.value = texte
    c.font = Font(bold=True, size=12, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22


def style_header(ws, row, headers, bg=NAVY, fg=WHITE):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color=fg, size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 28


def kpi_card(ws, row, col, label, valeur, unite, couleur):
    """Cree une carte KPI moderne sur 2 lignes x 1 colonne."""
    # Ligne valeur
    c1 = ws.cell(row=row, column=col, value=valeur)
    c1.font = Font(bold=True, size=22, color=couleur)
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c1.fill = PatternFill("solid", fgColor=GREY_LIGHT)
    ws.row_dimensions[row].height = 38

    # Ligne label
    c2 = ws.cell(row=row + 1, column=col, value=label + " (" + unite + ")")
    c2.font = Font(bold=True, size=10, color=NAVY)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    c2.fill = PatternFill("solid", fgColor=GREY_LIGHT)
    ws.row_dimensions[row + 1].height = 22


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# GABARIT 1 : DASHBOARD PROJET
# ============================================================
def gabarit_dashboard(wb):
    ws = wb.create_sheet("1. Dashboard")

    style_titre(ws, 'A1:H1', "DASHBOARD PROJET BIM", bg=NAVY, taille=18, hauteur=40)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"
    ws['E2'] = "Phase :"
    ws['E2'].font = Font(bold=True, color=NAVY)
    ws['F2'] = "DCE"
    ws['A3'] = "Date :"
    ws['A3'].font = Font(bold=True, color=NAVY)
    ws['B3'] = "[JJ/MM/AAAA]"
    ws['E3'] = "Maitre d'oeuvre :"
    ws['E3'].font = Font(bold=True, color=NAVY)
    ws['F3'] = "[A remplir]"

    # Ligne KPI
    style_sous_titre(ws, 5, 'A5:H5', "INDICATEURS CLES DU PROJET", bg=ORANGE)

    kpi_card(ws, 7, 1, "Surface SHON", 0, "m²", NAVY)
    kpi_card(ws, 7, 2, "Volume beton", 0, "m³", ORANGE)
    kpi_card(ws, 7, 3, "Nb elements", 0, "u", TURQUOISE)
    kpi_card(ws, 7, 4, "Cout prev.", 0, "k€", GREEN)
    kpi_card(ws, 7, 5, "IC carbone", 0, "kgCO2/m²", RED)
    kpi_card(ws, 7, 6, "Delai", 0, "mois", NAVY)
    kpi_card(ws, 7, 7, "Nb lots", 12, "", ORANGE)
    kpi_card(ws, 7, 8, "Conformite", 0, "%", GREEN)

    # Tableau repartition lots
    style_sous_titre(ws, 11, 'A11:H11', "REPARTITION PAR LOT (€ HT)", bg=TURQUOISE)
    style_header(ws, 13, ["Code", "Lot", "Montant HT (€)", "% Total",
                           "Avancement %", "", "", ""])

    lots = [
        ("01", "Terrassement"), ("02", "Gros oeuvre"),
        ("03", "Charpente / Couverture"), ("04", "Menuiseries ext."),
        ("05", "Cloisons / Doublages"), ("06", "Revetements sols"),
        ("07", "Menuiseries int."), ("08", "Peinture"),
        ("09", "Plomberie"), ("10", "Electricite"),
        ("11", "CVC"), ("12", "VRD"),
    ]

    for i, (code, nom) in enumerate(lots):
        r = 14 + i
        ws.cell(row=r, column=1, value=code).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=nom)
        ws.cell(row=r, column=3, value=0).number_format = '#,##0 €'
        ws.cell(row=r, column=4, value="=IF(SUM($C$14:$C$25)=0,0,C" + str(r) + "/SUM($C$14:$C$25))")
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=5, value=0).number_format = '0%'
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border()

    # Total
    r = 26
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=r, column=3, value="=SUM(C14:C25)").font = Font(bold=True, size=12)
    ws.cell(row=r, column=3).number_format = '#,##0 €'
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[r].height = 28

    # Data bar avancement
    ws.conditional_formatting.add("E14:E25",
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=100, color=TURQUOISE))

    # Color scale % lot
    ws.conditional_formatting.add("D14:D25",
        ColorScaleRule(start_type='min', start_color="FFFFFF",
                       end_type='max', end_color=ORANGE))

    # Graphique donut lots
    chart = DoughnutChart()
    chart.title = "Repartition budget par lot"
    labels = Reference(ws, min_col=2, min_row=14, max_row=25)
    data = Reference(ws, min_col=3, min_row=13, max_row=25)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 10
    chart.width = 15
    chart.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(chart, "J5")

    set_widths(ws, [10, 28, 18, 12, 15, 5, 5, 5])


# ============================================================
# GABARIT 2 : DPGF PROFESSIONNEL
# ============================================================
def gabarit_dpgf(wb):
    ws = wb.create_sheet("2. DPGF")

    style_titre(ws, 'A1:F1',
                "DECOMPOSITION DU PRIX GLOBAL ET FORFAITAIRE (DPGF)",
                bg=NAVY, taille=15)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"
    ws['A3'] = "Phase :"
    ws['A3'].font = Font(bold=True, color=NAVY)
    ws['B3'] = "DCE"

    style_header(ws, 5, ["Code", "Designation", "Unite", "Qte",
                          "PU HT (€)", "Total HT (€)"])

    lots_data = [
        ("LOT 01 - TERRASSEMENT", "8B5CF6", [
            ("01.01", "Decapage terre vegetale", "m²"),
            ("01.02", "Fouilles en pleine masse", "m³"),
            ("01.03", "Remblai compacte", "m³"),
        ]),
        ("LOT 02 - GROS OEUVRE", "3B82F6", [
            ("02.01", "Fondations semelles BA", "m³"),
            ("02.02", "Voiles BA", "m³"),
            ("02.03", "Dalles BA", "m³"),
            ("02.04", "Poteaux BA", "m³"),
            ("02.05", "Poutres BA", "m³"),
            ("02.06", "Maconnerie agglos", "m²"),
        ]),
        ("LOT 03 - CHARPENTE / COUVERTURE", "F59E0B", [
            ("03.01", "Charpente", "m²"),
            ("03.02", "Couverture", "m²"),
            ("03.03", "Zinguerie", "ml"),
        ]),
        ("LOT 04 - MENUISERIES EXTERIEURES", "22C55E", [
            ("04.01", "Fenetres", "U"),
            ("04.02", "Portes exterieures", "U"),
            ("04.03", "Occultations", "U"),
        ]),
        ("LOT 05 - CLOISONS / DOUBLAGES", "8B5CF6", [
            ("05.01", "Cloisons placo", "m²"),
            ("05.02", "Doublages isolants", "m²"),
            ("05.03", "Faux plafonds", "m²"),
        ]),
        ("LOT 06 - REVETEMENTS SOLS", "14B8A6", [
            ("06.01", "Chape / ragreage", "m²"),
            ("06.02", "Carrelage", "m²"),
            ("06.03", "Parquet / stratifie", "m²"),
        ]),
        ("LOT 07 - MENUISERIES INTERIEURES", "F59E0B", [
            ("07.01", "Portes interieures", "U"),
            ("07.02", "Placards", "ml"),
        ]),
        ("LOT 08 - PEINTURE", "EF4444", [
            ("08.01", "Peinture murs/plafonds", "m²"),
            ("08.02", "Peinture boiseries", "m²"),
        ]),
        ("LOT 09 - PLOMBERIE", "3B82F6", [
            ("09.01", "Alimentation EF/ECS", "Ft"),
            ("09.02", "Evacuations", "Ft"),
            ("09.03", "Appareils sanitaires", "Ens"),
        ]),
        ("LOT 10 - ELECTRICITE", "F59E0B", [
            ("10.01", "Tableau electrique", "U"),
            ("10.02", "Points lumineux", "U"),
            ("10.03", "Prises", "U"),
        ]),
        ("LOT 11 - CVC", "EF4444", [
            ("11.01", "Chauffage", "Ens"),
            ("11.02", "VMC", "Ens"),
        ]),
        ("LOT 12 - VRD", "22C55E", [
            ("12.01", "Voiries", "m²"),
            ("12.02", "Reseaux exterieurs", "Ft"),
        ]),
    ]

    row = 6
    first_data_row = None
    for nom_lot, couleur, items in lots_data:
        style_sous_titre(ws, row, "A" + str(row) + ":F" + str(row),
                         nom_lot, bg=couleur)
        row += 1
        for code, designation, unite in items:
            if first_data_row is None:
                first_data_row = row
            ws.cell(row=row, column=1, value=code).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=designation)
            ws.cell(row=row, column=3, value=unite).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=0)
            ws.cell(row=row, column=5, value=0).number_format = '#,##0.00 €'
            ws.cell(row=row, column=6,
                    value="=D" + str(row) + "*E" + str(row)).number_format = '#,##0.00 €'
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border()
            row += 1

    # Totaux HT / TVA / TTC
    last_data_row = row - 1
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="TOTAL HT").font = Font(bold=True, size=13, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', indent=2, vertical='center')
    total_ht = ws.cell(row=row, column=6,
                       value="=SUM(F" + str(first_data_row) + ":F" + str(last_data_row) + ")")
    total_ht.font = Font(bold=True, size=13)
    total_ht.number_format = '#,##0.00 €'
    total_ht.fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[row].height = 28
    ht_row = row
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="TVA 20%").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', indent=2)
    ws.cell(row=row, column=6, value="=F" + str(ht_row) + "*0.2").number_format = '#,##0.00 €'
    ws.cell(row=row, column=6).font = Font(bold=True, size=12)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="TOTAL TTC").font = Font(bold=True, size=14, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ORANGE)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', indent=2, vertical='center')
    ws.cell(row=row, column=6, value="=F" + str(ht_row) + "*1.2").number_format = '#,##0.00 €'
    ws.cell(row=row, column=6).font = Font(bold=True, size=14)
    ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[row].height = 32

    set_widths(ws, [10, 50, 10, 12, 15, 18])


# ============================================================
# GABARIT 3 : PLANNING CHANTIER
# ============================================================
def gabarit_planning(wb):
    ws = wb.create_sheet("3. Planning")

    style_titre(ws, 'A1:P1', "PLANNING CHANTIER (Gantt simplifie)",
                bg=NAVY, taille=15)

    ws['A2'] = "Debut :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[JJ/MM/AAAA]"
    ws['E2'] = "Duree totale :"
    ws['E2'].font = Font(bold=True, color=NAVY)
    ws['F2'] = 12
    ws['G2'] = "mois"

    # En-tete tableau : tache + 12 mois
    headers = ["Tache", "Debut (S)", "Duree (S)"]
    for i in range(1, 13):
        headers.append("M" + str(i))
    style_header(ws, 4, headers)

    taches = [
        "Installation chantier",
        "Terrassement",
        "Fondations",
        "Gros oeuvre - RdC",
        "Gros oeuvre - Etages",
        "Charpente / Couverture",
        "Menuiseries ext.",
        "Cloisons / Doublages",
        "Plomberie / Electricite / CVC",
        "Revetements sols",
        "Menuiseries int. / Peinture",
        "VRD / Exterieurs",
        "Reception / OPR",
    ]

    fill_orange = PatternFill("solid", fgColor=ORANGE)

    for i, t in enumerate(taches):
        r = 5 + i
        ws.cell(row=r, column=1, value=t)
        ws.cell(row=r, column=2, value=0)  # Debut en semaines
        ws.cell(row=r, column=3, value=0)  # Duree en semaines
        for col in range(1, 16):
            ws.cell(row=r, column=col).border = thin_border()
            if col >= 4:
                # Formule : colore la cellule si le mois est dans la plage [debut, debut+duree]
                # Mois n = semaines [4*(n-1)+1 .. 4*n]
                mois = col - 3
                debut_col = "$B" + str(r)
                duree_col = "$C" + str(r)
                # Rempli si (debut < 4*mois) ET (debut + duree > 4*(mois-1))
                formule = ('=IF(AND(' + debut_col + '<' + str(4 * mois) +
                           ',' + debut_col + '+' + duree_col + '>' + str(4 * (mois - 1)) +
                           '),"X","")')
                c = ws.cell(row=r, column=col, value=formule)
                c.alignment = Alignment(horizontal='center')
                c.font = Font(color=ORANGE, bold=True)

        ws.row_dimensions[r].height = 22

    # Format conditionnel : colore si "X"
    ws.conditional_formatting.add("D5:O" + str(4 + len(taches)),
        CellIsRule(operator='equal', formula=['"X"'], fill=fill_orange,
                   font=Font(color=ORANGE, bold=True)))

    set_widths(ws, [35, 10, 10] + [6] * 12 + [5])


# ============================================================
# GABARIT 4 : BILAN CARBONE RE2020
# ============================================================
def gabarit_bilan_carbone(wb):
    ws = wb.create_sheet("4. Bilan carbone")

    style_titre(ws, 'A1:F1', "BILAN CARBONE RE2020 - IC Construction",
                bg=GREEN, taille=15)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"
    ws['A3'] = "SHON :"
    ws['A3'].font = Font(bold=True, color=NAVY)
    ws['B3'] = 100
    ws['C3'] = "m²"

    # KPI cards
    style_sous_titre(ws, 5, 'A5:F5', "INDICATEURS CARBONE", bg=ORANGE)

    kpi_card(ws, 7, 1, "IC total", "=F50", "kgCO2eq", RED)
    kpi_card(ws, 7, 2, "IC / m²", "=F50/B3", "kgCO2eq/m²", ORANGE)
    kpi_card(ws, 7, 3, "Seuil 2025", 530, "kgCO2eq/m²", NAVY)
    kpi_card(ws, 7, 4, "Ecart seuil", "=F50/B3-530", "kgCO2eq/m²", NAVY)
    kpi_card(ws, 7, 5, "Conformite", '=IF(F50/B3<=530,"OK","DEPASSE")', "", GREEN)

    # Tableau lots carbone
    style_sous_titre(ws, 11, 'A11:F11', "DETAIL PAR LOT", bg=TURQUOISE)
    style_header(ws, 13, ["Lot", "Ouvrage", "Qte", "Unite",
                           "Facteur (kgCO2/U)", "kgCO2eq"])

    FDES = {
        "Beton": 250, "Acier": 2.0, "Bois charpente": -25,
        "Couverture": 18, "PVC menuiserie": 90, "Placo": 3,
        "Laine verre": 6, "PSE": 12,
    }

    lignes = [
        ("GROS OEUVRE", "Beton fondations", "m³", FDES["Beton"]),
        ("GROS OEUVRE", "Beton murs/voiles", "m³", FDES["Beton"]),
        ("GROS OEUVRE", "Beton dalles", "m³", FDES["Beton"]),
        ("GROS OEUVRE", "Beton poteaux/poutres", "m³", FDES["Beton"]),
        ("GROS OEUVRE", "Acier BA", "kg", FDES["Acier"]),
        ("CHARPENTE", "Charpente bois", "m²", FDES["Bois charpente"]),
        ("CHARPENTE", "Couverture", "m²", FDES["Couverture"]),
        ("MENUISERIES", "Fenetres PVC", "m²", FDES["PVC menuiserie"]),
        ("CLOISONS", "Cloisons placo", "m²", FDES["Placo"]),
        ("ISOLATION", "Laine de verre", "m²", FDES["Laine verre"]),
        ("ISOLATION", "PSE facade", "m²", FDES["PSE"]),
    ]

    row = 14
    for lot, ouvrage, unite, facteur in lignes:
        ws.cell(row=row, column=1, value=lot)
        ws.cell(row=row, column=2, value=ouvrage)
        ws.cell(row=row, column=3, value=0)  # Quantite a remplir
        ws.cell(row=row, column=4, value=unite).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=facteur)
        ws.cell(row=row, column=6, value="=C" + str(row) + "*E" + str(row))
        ws.cell(row=row, column=6).number_format = '#,##0'
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1

    # TOTAL (ligne 50 pour les KPI)
    while row < 50:
        row += 1

    ws.merge_cells('A50:E50')
    ws['A50'] = "TOTAL IC CONSTRUCTION (kgCO2eq)"
    ws['A50'].font = Font(bold=True, size=13, color=WHITE)
    ws['A50'].fill = PatternFill("solid", fgColor=RED)
    ws['A50'].alignment = Alignment(horizontal='right', indent=2, vertical='center')
    ws['F50'] = "=SUM(F14:F24)"
    ws['F50'].font = Font(bold=True, size=13)
    ws['F50'].number_format = '#,##0'
    ws['F50'].fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[50].height = 30

    # Tableau seuils RE2020
    style_sous_titre(ws, 52, 'A52:F52', "SEUILS RE2020 (logement)", bg=NAVY)
    style_header(ws, 53, ["Periode", "Seuil (kgCO2/m²)", "Projet (kgCO2/m²)",
                           "Ecart", "Statut", ""])

    seuils = [
        ("2022-2024", 640), ("2025-2027", 530),
        ("2028-2030", 475), ("2031+", 415),
    ]
    for i, (periode, seuil) in enumerate(seuils):
        r = 54 + i
        ws.cell(row=r, column=1, value=periode)
        ws.cell(row=r, column=2, value=seuil)
        ws.cell(row=r, column=3, value="=$F$50/$B$3").number_format = '0'
        ws.cell(row=r, column=4, value="=C" + str(r) + "-B" + str(r)).number_format = '+0;-0;0'
        ws.cell(row=r, column=5,
                value='=IF(C' + str(r) + '<=B' + str(r) + ',"Respecte","Depasse")')
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border()
            ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')

    # Format conditionnel statut
    ws.conditional_formatting.add("E54:E57",
        CellIsRule(operator='equal', formula=['"Respecte"'],
                   fill=PatternFill("solid", fgColor="DCFCE7"),
                   font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add("E54:E57",
        CellIsRule(operator='equal', formula=['"Depasse"'],
                   fill=PatternFill("solid", fgColor="FEE2E2"),
                   font=Font(color=RED, bold=True)))

    # Graphique barres seuils
    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Comparaison seuils RE2020"
    chart.y_axis.title = 'kgCO2eq/m²'
    data = Reference(ws, min_col=2, min_row=53, max_col=3, max_row=57)
    cats = Reference(ws, min_col=1, min_row=54, max_row=57)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "H11")

    set_widths(ws, [18, 28, 12, 10, 18, 15])


# ============================================================
# GABARIT 5 : CONTROLE QUALITE MAQUETTE BIM
# ============================================================
def gabarit_controle_qualite(wb):
    ws = wb.create_sheet("5. Controle BIM")

    style_titre(ws, 'A1:F1', "CONTROLE QUALITE MAQUETTE BIM",
                bg=TURQUOISE, fg=WHITE, taille=15)

    ws['A2'] = "Maquette :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[fichier.ifc]"
    ws['A3'] = "Controleur :"
    ws['A3'].font = Font(bold=True, color=NAVY)
    ws['B3'] = "[nom]"

    # KPI score global
    style_sous_titre(ws, 5, 'A5:F5', "SCORE QUALITE", bg=ORANGE)

    kpi_card(ws, 7, 1, "Score total", "=COUNTIF(D13:D50,\"OK\")", "/ 30", GREEN)
    kpi_card(ws, 7, 2, "Non conformes", "=COUNTIF(D13:D50,\"KO\")", "", RED)
    kpi_card(ws, 7, 3, "A verifier", "=COUNTIF(D13:D50,\"?\")", "", YELLOW)
    kpi_card(ws, 7, 4, "Conformite",
             "=IFERROR(COUNTIF(D13:D50,\"OK\")/(COUNTA(D13:D50)),0)", "%", NAVY)

    # Checklist
    style_sous_titre(ws, 11, 'A11:F11', "CHECKLIST CONTROLE", bg=NAVY)
    style_header(ws, 12, ["Categorie", "Point de controle", "Critere attendu",
                           "Statut", "Commentaire", ""])

    controles = [
        ("Structure IFC", "Fichier IFC valide", "Ouverture sans erreur"),
        ("Structure IFC", "Version IFC", "IFC2x3 ou IFC4"),
        ("Structure IFC", "Unites declarees", "Metres, m², m³"),
        ("Structure IFC", "Projet/Site/Batiment/Etages", "Hierarchie complete"),
        ("Geometrie", "Murs presents", "IfcWall non vide"),
        ("Geometrie", "Dalles presentes", "IfcSlab non vide"),
        ("Geometrie", "Poteaux/Poutres", "IfcColumn/IfcBeam"),
        ("Geometrie", "Fondations", "IfcFooting/IfcPile"),
        ("Geometrie", "Menuiseries", "IfcDoor/IfcWindow"),
        ("Geometrie", "Toiture", "IfcRoof"),
        ("Quantites", "Volumes renseignes", "NetVolume > 0"),
        ("Quantites", "Surfaces renseignees", "NetArea > 0"),
        ("Quantites", "Longueurs renseignees", "Length > 0"),
        ("Psets", "Pset_WallCommon", "LoadBearing, IsExternal"),
        ("Psets", "Pset_SlabCommon", "LoadBearing, PitchAngle"),
        ("Psets", "Pset_DoorCommon", "FireRating, IsExternal"),
        ("Psets", "Pset_WindowCommon", "GlazingAreaFraction"),
        ("Materiaux", "Materiau beton defini", "IfcMaterial assigne"),
        ("Materiaux", "Materiau isolant defini", "IfcMaterial assigne"),
        ("Materiaux", "Materiau menuiserie", "IfcMaterial assigne"),
        ("Classification", "Code Uniformat/Omniclass", "Optionnel"),
        ("Classification", "Identifiants uniques", "GlobalId present"),
        ("Nommage", "Noms coherents", "Mur_RDC_01, etc."),
        ("Nommage", "Types definis", "IfcWallType, etc."),
        ("Etages", "Noms etages explicites", "RDC, R+1, etc."),
        ("Etages", "Elements affectes etage", "ContainedInStructure"),
        ("Coherence", "Pas de doublons", "GUID uniques"),
        ("Coherence", "Elements orphelins", "Tous relies au batiment"),
        ("Livrables", "Nomenclature quantites", "Excel/BCF"),
        ("Livrables", "Rapport de controle", "PDF/Excel"),
    ]

    for i, (cat, point, critere) in enumerate(controles):
        r = 13 + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=point)
        ws.cell(row=r, column=3, value=critere)
        ws.cell(row=r, column=4, value="").alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5, value="")
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border()
        ws.row_dimensions[r].height = 20

    # Format conditionnel statut
    statut_range = "D13:D" + str(12 + len(controles))
    ws.conditional_formatting.add(statut_range,
        CellIsRule(operator='equal', formula=['"OK"'],
                   fill=PatternFill("solid", fgColor="DCFCE7"),
                   font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add(statut_range,
        CellIsRule(operator='equal', formula=['"KO"'],
                   fill=PatternFill("solid", fgColor="FEE2E2"),
                   font=Font(color=RED, bold=True)))
    ws.conditional_formatting.add(statut_range,
        CellIsRule(operator='equal', formula=['"?"'],
                   fill=PatternFill("solid", fgColor="FEF3C7"),
                   font=Font(color=YELLOW, bold=True)))

    # Validation donnees : OK / KO / ?
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"OK,KO,?"', allow_blank=True)
    dv.add(statut_range)
    ws.add_data_validation(dv)

    set_widths(ws, [18, 32, 32, 12, 35, 5])


# ============================================================
# GABARIT 6 : METRE DETAILLE PAR OUVRAGE
# ============================================================
def gabarit_metre(wb):
    ws = wb.create_sheet("6. Metre")

    style_titre(ws, 'A1:H1', "METRE DETAILLE PAR OUVRAGE", bg=NAVY, taille=15)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"

    style_sous_titre(ws, 4, 'A4:H4',
                     "CALCULS DE METRES (longueur x largeur x hauteur)", bg=ORANGE)

    style_header(ws, 6, ["Code", "Ouvrage", "Local / Zone",
                          "L (m)", "l (m)", "H (m)", "Nb", "Total"])

    ouvrages = [
        ("M.01", "Mur exterieur BA 20cm", "Facade Nord"),
        ("M.02", "Mur exterieur BA 20cm", "Facade Sud"),
        ("M.03", "Mur exterieur BA 20cm", "Facade Est"),
        ("M.04", "Mur exterieur BA 20cm", "Facade Ouest"),
        ("M.05", "Mur interieur porteur", "Refend central"),
        ("C.01", "Cloison placo 72/48", "Sejour/Cuisine"),
        ("C.02", "Cloison placo 72/48", "Chambres"),
        ("D.01", "Dalle BA 20cm", "RdC"),
        ("D.02", "Dalle BA 20cm", "R+1"),
        ("D.03", "Dalle BA 20cm", "Toiture terrasse"),
        ("P.01", "Poteau BA 25x25", "Structure"),
        ("P.02", "Poutre BA 20x40", "Linteaux"),
        ("F.01", "Semelle filante", "Peripherie"),
        ("F.02", "Semelle isolee", "Sous poteaux"),
        ("M.10", "Menuiserie fenetre", "Baies vitrees"),
        ("M.11", "Menuiserie porte ext.", "Entree"),
        ("T.01", "Toiture couverture", "Pente sud"),
        ("T.02", "Toiture couverture", "Pente nord"),
        ("R.01", "Revetement sol carrelage", "Pieces humides"),
        ("R.02", "Revetement sol parquet", "Pieces seches"),
    ]

    for i, (code, ouvrage, zone) in enumerate(ouvrages):
        r = 7 + i
        ws.cell(row=r, column=1, value=code).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=ouvrage)
        ws.cell(row=r, column=3, value=zone)
        ws.cell(row=r, column=4, value=0)
        ws.cell(row=r, column=5, value=0)
        ws.cell(row=r, column=6, value=0)
        ws.cell(row=r, column=7, value=1)
        ws.cell(row=r, column=8,
                value="=D" + str(r) + "*E" + str(r) + "*F" + str(r) + "*G" + str(r))
        ws.cell(row=r, column=8).number_format = '0.00'
        ws.cell(row=r, column=8).font = Font(bold=True, color=ORANGE)
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = thin_border()

    # TOTAL
    r = 7 + len(ouvrages) + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="TOTAL METRE").font = Font(bold=True, size=12, color=WHITE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='right', indent=2, vertical='center')
    ws.cell(row=r, column=8,
            value="=SUM(H7:H" + str(6 + len(ouvrages)) + ")").font = Font(bold=True, size=12)
    ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[r].height = 26

    set_widths(ws, [8, 28, 22, 10, 10, 10, 8, 12])


# ============================================================
# GABARIT 7 : DQE - DEVIS QUANTITATIF ESTIMATIF
# ============================================================
def gabarit_dqe(wb):
    ws = wb.create_sheet("7. DQE")

    style_titre(ws, 'A1:G1', "DEVIS QUANTITATIF ESTIMATIF (DQE)",
                bg=NAVY, taille=15)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"
    ws['D2'] = "Date :"
    ws['D2'].font = Font(bold=True, color=NAVY)
    ws['E2'] = "[JJ/MM/AAAA]"

    style_header(ws, 4, ["Code", "Designation", "Unite", "Qte",
                          "PU mat. (€)", "PU pose (€)", "Total (€)"])

    postes = [
        ("01.01", "Terrassement general", "m³"),
        ("02.01", "Beton fondations C25/30", "m³"),
        ("02.02", "Beton voiles C25/30", "m³"),
        ("02.03", "Beton dalles C25/30", "m³"),
        ("02.04", "Acier BA HA500", "kg"),
        ("02.05", "Maconnerie parpaing 20cm", "m²"),
        ("03.01", "Charpente bois", "m³"),
        ("03.02", "Couverture tuiles", "m²"),
        ("04.01", "Menuiserie PVC double vitrage", "m²"),
        ("04.02", "Porte bois entree", "U"),
        ("05.01", "Cloison placo BA13", "m²"),
        ("05.02", "Isolation laine de verre 100mm", "m²"),
        ("06.01", "Carrelage gres cerame", "m²"),
        ("06.02", "Parquet stratifie", "m²"),
        ("08.01", "Peinture acrylique 2c", "m²"),
        ("09.01", "Reseau eau froide PER", "ml"),
        ("10.01", "Cable electrique", "ml"),
        ("11.01", "VMC simple flux", "U"),
        ("12.01", "Enrobe voirie", "m²"),
    ]

    row = 5
    for code, design, unite in postes:
        ws.cell(row=row, column=1, value=code).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=design)
        ws.cell(row=row, column=3, value=unite).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=0)
        ws.cell(row=row, column=5, value=0).number_format = '#,##0.00 €'
        ws.cell(row=row, column=6, value=0).number_format = '#,##0.00 €'
        ws.cell(row=row, column=7,
                value="=D" + str(row) + "*(E" + str(row) + "+F" + str(row) + ")")
        ws.cell(row=row, column=7).number_format = '#,##0.00 €'
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1

    # Totaux
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="TOTAL HT").font = Font(bold=True, size=13, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', indent=2, vertical='center')
    ws.cell(row=row, column=7,
            value="=SUM(G5:G" + str(row - 2) + ")").font = Font(bold=True, size=13)
    ws.cell(row=row, column=7).number_format = '#,##0.00 €'
    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[row].height = 28

    set_widths(ws, [10, 42, 10, 10, 15, 15, 18])


# ============================================================
# GABARIT 8 : ETUDE DE PRIX - SOUS-DETAIL
# ============================================================
def gabarit_sous_detail(wb):
    ws = wb.create_sheet("8. Sous-detail")

    style_titre(ws, 'A1:F1', "ETUDE DE PRIX - SOUS-DETAIL D'OUVRAGE",
                bg=NAVY, taille=15)

    ws['A2'] = "Ouvrage :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[Ex : m³ beton arme voile]"
    ws['A3'] = "Unite :"
    ws['A3'].font = Font(bold=True, color=NAVY)
    ws['B3'] = "m³"

    # MATERIAUX
    style_sous_titre(ws, 5, 'A5:F5', "DEBOURSE MATERIAUX", bg=ORANGE)
    style_header(ws, 6, ["Composant", "Qte / U", "Unite", "Prix unitaire",
                          "Montant", "% debourse"])

    mat = [
        ("Beton C25/30 pret a l'emploi", 1.05, "m³"),
        ("Acier HA500 (150 kg/m³)", 150, "kg"),
        ("Coffrage ordinaire", 10, "m²"),
        ("Pompage beton", 1, "m³"),
        ("Adjuvants / produits divers", 1, "Ft"),
    ]
    row = 7
    first_mat = row
    for comp, qte, u in mat:
        ws.cell(row=row, column=1, value=comp)
        ws.cell(row=row, column=2, value=qte)
        ws.cell(row=row, column=3, value=u).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=0).number_format = '#,##0.00 €'
        ws.cell(row=row, column=5,
                value="=B" + str(row) + "*D" + str(row)).number_format = '#,##0.00 €'
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1
    last_mat = row - 1

    # MAIN D'OEUVRE
    row += 1
    style_sous_titre(ws, row, 'A' + str(row) + ':F' + str(row),
                     "DEBOURSE MAIN D'OEUVRE", bg=ORANGE)
    row += 1
    style_header(ws, row, ["Categorie", "Heures / U", "Unite",
                            "Taux horaire", "Montant", ""])
    row += 1
    mo = [
        ("Chef d'equipe N4", 1.0, "h"),
        ("Ouvrier qualifie N3", 3.5, "h"),
        ("Manoeuvre N1", 2.0, "h"),
    ]
    first_mo = row
    for cat, h, u in mo:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=h)
        ws.cell(row=row, column=3, value=u).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=0).number_format = '#,##0.00 €'
        ws.cell(row=row, column=5,
                value="=B" + str(row) + "*D" + str(row)).number_format = '#,##0.00 €'
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1
    last_mo = row - 1

    # MATERIEL
    row += 1
    style_sous_titre(ws, row, 'A' + str(row) + ':F' + str(row),
                     "DEBOURSE MATERIEL", bg=ORANGE)
    row += 1
    style_header(ws, row, ["Equipement", "Duree", "Unite",
                            "Cout unitaire", "Montant", ""])
    row += 1
    mat_eq = [
        ("Grue de chantier (prorata)", 0.5, "h"),
        ("Outillage divers (prorata)", 1, "Ft"),
    ]
    first_eq = row
    for eq, d, u in mat_eq:
        ws.cell(row=row, column=1, value=eq)
        ws.cell(row=row, column=2, value=d)
        ws.cell(row=row, column=3, value=u).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=0).number_format = '#,##0.00 €'
        ws.cell(row=row, column=5,
                value="=B" + str(row) + "*D" + str(row)).number_format = '#,##0.00 €'
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1
    last_eq = row - 1

    # SYNTHESE
    row += 2
    ws.cell(row=row, column=1, value="Debourse sec materiaux").font = Font(bold=True)
    ws.cell(row=row, column=5,
            value="=SUM(E" + str(first_mat) + ":E" + str(last_mat) + ")").number_format = '#,##0.00 €'
    row += 1
    ws.cell(row=row, column=1, value="Debourse sec main d'oeuvre").font = Font(bold=True)
    ws.cell(row=row, column=5,
            value="=SUM(E" + str(first_mo) + ":E" + str(last_mo) + ")").number_format = '#,##0.00 €'
    row += 1
    ws.cell(row=row, column=1, value="Debourse sec materiel").font = Font(bold=True)
    ws.cell(row=row, column=5,
            value="=SUM(E" + str(first_eq) + ":E" + str(last_eq) + ")").number_format = '#,##0.00 €'
    ds_row = row
    row += 1
    ws.cell(row=row, column=1, value="DEBOURSE SEC TOTAL").font = Font(bold=True, color=NAVY)
    ws.cell(row=row, column=5,
            value="=SUM(E" + str(ds_row - 2) + ":E" + str(ds_row) + ")")
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).number_format = '#,##0.00 €'
    ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=GREY_LIGHT)
    ds_tot = row
    row += 1
    ws.cell(row=row, column=1, value="Frais de chantier (10%)")
    ws.cell(row=row, column=5, value="=E" + str(ds_tot) + "*0.1").number_format = '#,##0.00 €'
    row += 1
    ws.cell(row=row, column=1, value="Frais generaux (12%)")
    ws.cell(row=row, column=5, value="=E" + str(ds_tot) + "*0.12").number_format = '#,##0.00 €'
    row += 1
    ws.cell(row=row, column=1, value="Benefice et alea (8%)")
    ws.cell(row=row, column=5, value="=E" + str(ds_tot) + "*0.08").number_format = '#,##0.00 €'
    row += 1
    ws.cell(row=row, column=1, value="PRIX DE VENTE UNITAIRE HT").font = Font(bold=True, size=13, color=WHITE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ORANGE)
    ws.cell(row=row, column=5,
            value="=E" + str(ds_tot) + "*1.3").number_format = '#,##0.00 €'
    ws.cell(row=row, column=5).font = Font(bold=True, size=13)
    ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=YELLOW)
    ws.row_dimensions[row].height = 28

    set_widths(ws, [35, 12, 10, 15, 18, 12])


# ============================================================
# GABARIT 9 : NOTE DE SYNTHESE PROJET
# ============================================================
def gabarit_note_synthese(wb):
    ws = wb.create_sheet("9. Note synthese")

    style_titre(ws, 'A1:D1', "NOTE DE SYNTHESE - PROJET BIM",
                bg=NAVY, taille=15)

    ws['A3'] = "Projet :"
    ws['A3'].font = Font(bold=True, color=NAVY)
    ws['B3'] = "[A remplir]"
    ws['A4'] = "Maitre d'ouvrage :"
    ws['A4'].font = Font(bold=True, color=NAVY)
    ws['B4'] = "[A remplir]"
    ws['A5'] = "Adresse :"
    ws['A5'].font = Font(bold=True, color=NAVY)
    ws['B5'] = "[A remplir]"
    ws['A6'] = "Phase :"
    ws['A6'].font = Font(bold=True, color=NAVY)
    ws['B6'] = "DCE / EXE / APS / APD"

    sections = [
        ("1. CONTEXTE DU PROJET",
         "Presentation generale du projet, type d'ouvrage, destination, "
         "contraintes reglementaires (PLU, ERP, RE2020...)."),
        ("2. CARACTERISTIQUES TECHNIQUES",
         "Surface SHON / SHOB, nombre de niveaux, type de structure, "
         "procede constructif, systemes techniques (CVC, elec, plomb...)."),
        ("3. METHODOLOGIE BIM",
         "Logiciels utilises (Revit, ArchiCAD, eveBIM), LOD cible, "
         "convention BIM, plateforme collaborative, processus IFC."),
        ("4. QUANTITES EXTRAITES",
         "Synthese des metres extraits de la maquette : volumes beton, "
         "surfaces, menuiseries. Voir onglet 'Metre' pour detail."),
        ("5. ESTIMATION FINANCIERE",
         "Cout travaux HT par lot, ratio €/m², comparaison avec references "
         "du secteur. Voir onglets 'DPGF' et 'DQE'."),
        ("6. IMPACT ENVIRONNEMENTAL",
         "IC construction (kgCO2eq/m²), comparaison seuils RE2020, "
         "preconisations bas carbone. Voir onglet 'Bilan carbone'."),
        ("7. PLANNING PREVISIONNEL",
         "Duree totale, chemin critique, phases. Voir onglet 'Planning'."),
        ("8. CONTROLE QUALITE",
         "Niveau de conformite de la maquette, points de vigilance, "
         "actions correctives. Voir onglet 'Controle BIM'."),
        ("9. VARIANTES ETUDIEES",
         "Options techniques envisagees, analyse comparative (cout, delai, "
         "performance). Voir onglet 'Comparatif variantes'."),
        ("10. CONCLUSION",
         "Synthese des elements cles, faisabilite, recommandations, "
         "suites a donner."),
    ]

    row = 8
    for titre, contenu in sections:
        ws.merge_cells('A' + str(row) + ':D' + str(row))
        ws.cell(row=row, column=1, value=titre).font = Font(bold=True, size=12, color=WHITE)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ORANGE)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1, vertical='center')
        ws.row_dimensions[row].height = 24
        row += 1

        ws.merge_cells('A' + str(row) + ':D' + str(row + 3))
        c = ws.cell(row=row, column=1, value=contenu + "\n\n[Completer ici...]")
        c.alignment = Alignment(wrap_text=True, vertical='top', indent=1)
        c.font = Font(size=10)
        c.fill = PatternFill("solid", fgColor=GREY_LIGHT)
        ws.row_dimensions[row].height = 22
        for i in range(4):
            ws.row_dimensions[row + i].height = 22
        row += 5

    set_widths(ws, [25, 25, 25, 25])


# ============================================================
# GABARIT 10 : DESCENTE DE CHARGES SIMPLIFIEE
# ============================================================
def gabarit_descente_charges(wb):
    ws = wb.create_sheet("10. Charges")

    style_titre(ws, 'A1:G1', "DESCENTE DE CHARGES SIMPLIFIEE",
                bg=NAVY, taille=15)

    ws['A2'] = "Element :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[Ex : Poteau P1]"

    # Charges permanentes
    style_sous_titre(ws, 4, 'A4:G4', "CHARGES PERMANENTES (G)", bg=ORANGE)
    style_header(ws, 5, ["Niveau", "Element", "Poids vol. (kN/m³)",
                          "Surf/Long (m² ou ml)", "Epaisseur (m)",
                          "Charge (kN/m²)", "Total (kN)"])

    charges_g = [
        ("Toiture", "Couverture + charpente", 0, 1, 0, 1.5),
        ("R+1", "Dalle BA + revetement", 25, 1, 0.22, 0),
        ("R+1", "Cloisons legeres", 0, 1, 0, 0.5),
        ("RdC", "Dalle BA + revetement", 25, 1, 0.22, 0),
        ("RdC", "Cloisons legeres", 0, 1, 0, 0.5),
        ("Fondations", "Semelle BA", 25, 1, 0.4, 0),
    ]

    row = 6
    first_g = row
    for niv, elem, pv, surf, ep, q in charges_g:
        ws.cell(row=row, column=1, value=niv)
        ws.cell(row=row, column=2, value=elem)
        ws.cell(row=row, column=3, value=pv)
        ws.cell(row=row, column=4, value=surf)
        ws.cell(row=row, column=5, value=ep)
        ws.cell(row=row, column=6, value=q)
        # Total = surf * (q OU pv*ep)
        ws.cell(row=row, column=7,
                value="=D" + str(row) + "*IF(F" + str(row) + ">0,F" + str(row) +
                ",C" + str(row) + "*E" + str(row) + ")")
        ws.cell(row=row, column=7).number_format = '0.00'
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1
    last_g = row - 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL G (kN)").font = Font(bold=True, color=NAVY)
    ws.cell(row=row, column=7,
            value="=SUM(G" + str(first_g) + ":G" + str(last_g) + ")")
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=GREY_LIGHT)
    g_row = row

    # Charges exploitation
    row += 2
    style_sous_titre(ws, row, 'A' + str(row) + ':G' + str(row),
                     "CHARGES D'EXPLOITATION (Q)", bg=ORANGE)
    row += 1
    style_header(ws, row, ["Niveau", "Usage", "",
                            "Surface (m²)", "", "q (kN/m²)", "Total (kN)"])
    row += 1
    charges_q = [
        ("Toiture terrasse", "Non accessible", 1, 1.0),
        ("R+1", "Logement", 1, 1.5),
        ("RdC", "Logement", 1, 1.5),
    ]
    first_q = row
    for niv, us, surf, q in charges_q:
        ws.cell(row=row, column=1, value=niv)
        ws.cell(row=row, column=2, value=us)
        ws.cell(row=row, column=4, value=surf)
        ws.cell(row=row, column=6, value=q)
        ws.cell(row=row, column=7,
                value="=D" + str(row) + "*F" + str(row)).number_format = '0.00'
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border()
        row += 1
    last_q = row - 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL Q (kN)").font = Font(bold=True, color=NAVY)
    ws.cell(row=row, column=7,
            value="=SUM(G" + str(first_q) + ":G" + str(last_q) + ")")
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=GREY_LIGHT)
    q_row = row

    # Combinaisons
    row += 2
    style_sous_titre(ws, row, 'A' + str(row) + ':G' + str(row),
                     "COMBINAISONS ELU / ELS", bg=NAVY)
    row += 1
    ws.cell(row=row, column=1, value="ELU : 1.35 G + 1.5 Q").font = Font(bold=True)
    ws.cell(row=row, column=7,
            value="=1.35*G" + str(g_row) + "+1.5*G" + str(q_row))
    ws.cell(row=row, column=7).number_format = '0.00'
    ws.cell(row=row, column=7).font = Font(bold=True, size=12, color=RED)
    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=YELLOW)
    row += 1
    ws.cell(row=row, column=1, value="ELS : G + Q").font = Font(bold=True)
    ws.cell(row=row, column=7, value="=G" + str(g_row) + "+G" + str(q_row))
    ws.cell(row=row, column=7).number_format = '0.00'
    ws.cell(row=row, column=7).font = Font(bold=True, size=12, color=NAVY)
    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=GREY_LIGHT)

    set_widths(ws, [15, 28, 15, 12, 12, 12, 12])


# ============================================================
# GABARIT 11 : BILAN THERMIQUE RE2020
# ============================================================
def gabarit_bilan_thermique(wb):
    ws = wb.create_sheet("11. Thermique")

    style_titre(ws, 'A1:F1', "BILAN THERMIQUE RE2020",
                bg="4DC7C7", fg=WHITE, taille=15)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"
    ws['D2'] = "Zone climatique :"
    ws['D2'].font = Font(bold=True, color=NAVY)
    ws['E2'] = "H1 / H2 / H3"

    # Indicateurs RE2020
    style_sous_titre(ws, 4, 'A4:F4', "INDICATEURS RE2020", bg=ORANGE)

    kpi_card(ws, 6, 1, "Bbio", 0, "pts", NAVY)
    kpi_card(ws, 6, 2, "Cep", 0, "kWh/m²/an", TURQUOISE)
    kpi_card(ws, 6, 3, "Cep,nr", 0, "kWh/m²/an", ORANGE)
    kpi_card(ws, 6, 4, "Ic energie", 0, "kgCO2/m²", GREEN)
    kpi_card(ws, 6, 5, "Ic construction", 0, "kgCO2/m²", RED)
    kpi_card(ws, 6, 6, "DH", 0, "°C.h", NAVY)

    # Parois
    style_sous_titre(ws, 10, 'A10:F10',
                     "CARACTERISTIQUES THERMIQUES DES PAROIS", bg=TURQUOISE)
    style_header(ws, 11, ["Paroi", "Surface (m²)", "U (W/m²K)",
                           "R (m²K/W)", "Type", "UxS (W/K)"])

    parois = [
        ("Murs exterieurs", "Murs mitoyens", "Toiture",
         "Plancher bas", "Fenetres", "Portes exterieures"),
    ][0]
    u_cible = [0.20, 0.40, 0.15, 0.22, 1.30, 1.50]

    for i, (p, u) in enumerate(zip(parois, u_cible)):
        r = 12 + i
        ws.cell(row=r, column=1, value=p)
        ws.cell(row=r, column=2, value=0)
        ws.cell(row=r, column=3, value=u)
        ws.cell(row=r, column=4, value="=IF(C" + str(r) + ">0,1/C" + str(r) + ",0)")
        ws.cell(row=r, column=4).number_format = '0.00'
        ws.cell(row=r, column=5, value="Opaque" if i < 4 else "Vitre")
        ws.cell(row=r, column=6, value="=B" + str(r) + "*C" + str(r))
        ws.cell(row=r, column=6).number_format = '0.00'
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border()

    # Total deperditions
    r = 12 + len(parois) + 1
    ws.cell(row=r, column=1, value="Deperditions totales (W/K)").font = Font(bold=True)
    ws.cell(row=r, column=6,
            value="=SUM(F12:F" + str(11 + len(parois)) + ")")
    ws.cell(row=r, column=6).font = Font(bold=True)
    ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=YELLOW)

    # Seuils reglementaires
    style_sous_titre(ws, r + 2, 'A' + str(r + 2) + ':F' + str(r + 2),
                     "SEUILS RE2020 (logement collectif)", bg=NAVY)
    style_header(ws, r + 3, ["Indicateur", "Seuil RE2020", "Projet",
                              "Ecart", "Statut", ""])

    seuils = [
        ("Bbio max", 72),
        ("Cep,nr max (kWh/m²/an)", 55),
        ("Cep max (kWh/m²/an)", 75),
        ("Ic energie max (kgCO2/m²)", 560),
        ("Ic construction 2025 (kgCO2/m²)", 740),
        ("DH max (°C.h)", 1250),
    ]
    for i, (ind, s) in enumerate(seuils):
        rr = r + 4 + i
        ws.cell(row=rr, column=1, value=ind)
        ws.cell(row=rr, column=2, value=s)
        ws.cell(row=rr, column=3, value=0)
        ws.cell(row=rr, column=4, value="=C" + str(rr) + "-B" + str(rr))
        ws.cell(row=rr, column=5,
                value='=IF(C' + str(rr) + '<=B' + str(rr) + ',"OK","KO")')
        for col in range(1, 6):
            ws.cell(row=rr, column=col).border = thin_border()

    set_widths(ws, [28, 15, 15, 12, 14, 15])


# ============================================================
# GABARIT 12 : COMPARATIF VARIANTES
# ============================================================
def gabarit_variantes(wb):
    ws = wb.create_sheet("12. Variantes")

    style_titre(ws, 'A1:F1', "COMPARATIF VARIANTES TECHNIQUES",
                bg=NAVY, taille=15)

    ws['A2'] = "Choix etudie :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[Ex : Solution structure - Beton vs Bois vs Metal]"

    style_sous_titre(ws, 4, 'A4:F4', "TABLEAU COMPARATIF", bg=ORANGE)
    style_header(ws, 6, ["Critere", "Poids (1-5)", "Variante 1",
                           "Variante 2", "Variante 3", "Observations"])

    criteres = [
        ("Cout investissement (€)", 5),
        ("Cout exploitation 10 ans", 3),
        ("Delai execution (mois)", 4),
        ("Performance thermique (U)", 4),
        ("Impact carbone (kgCO2/m²)", 5),
        ("Complexite mise en oeuvre", 3),
        ("Durabilite / entretien", 4),
        ("Esthetique / architectural", 2),
        ("Disponibilite materiaux", 3),
        ("Confort acoustique", 3),
    ]

    # Variantes
    for i, (crit, poids) in enumerate(criteres):
        r = 7 + i
        ws.cell(row=r, column=1, value=crit)
        ws.cell(row=r, column=2, value=poids).alignment = Alignment(horizontal='center')
        # Notes 1-5 a remplir
        ws.cell(row=r, column=3, value=0).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4, value=0).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=5, value=0).alignment = Alignment(horizontal='center')
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border()

    # Notes ponderees
    r = 7 + len(criteres) + 1
    ws.cell(row=r, column=1, value="SCORE PONDERE").font = Font(bold=True, size=12, color=WHITE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=r, column=2, value="").fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=r, column=3,
            value="=SUMPRODUCT($B$7:$B$" + str(6 + len(criteres)) + ",C7:C" + str(6 + len(criteres)) + ")")
    ws.cell(row=r, column=4,
            value="=SUMPRODUCT($B$7:$B$" + str(6 + len(criteres)) + ",D7:D" + str(6 + len(criteres)) + ")")
    ws.cell(row=r, column=5,
            value="=SUMPRODUCT($B$7:$B$" + str(6 + len(criteres)) + ",E7:E" + str(6 + len(criteres)) + ")")
    for col in range(3, 6):
        ws.cell(row=r, column=col).font = Font(bold=True, size=14, color=NAVY)
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
    ws.row_dimensions[r].height = 28

    # Recommandation
    r += 2
    ws.merge_cells('A' + str(r) + ':F' + str(r))
    ws.cell(row=r, column=1, value="RECOMMANDATION").font = Font(bold=True, size=12, color=WHITE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=ORANGE)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 24
    r += 1
    ws.merge_cells('A' + str(r) + ':F' + str(r + 4))
    c = ws.cell(row=r, column=1, value="[Justification du choix retenu et motifs techniques/economiques/environnementaux...]")
    c.alignment = Alignment(wrap_text=True, vertical='top', indent=1)
    c.fill = PatternFill("solid", fgColor=GREY_LIGHT)
    for i in range(5):
        ws.row_dimensions[r + i].height = 22

    # Format conditionnel : mettre en vert la meilleure variante
    score_range = "C" + str(7 + len(criteres) + 1) + ":E" + str(7 + len(criteres) + 1)
    ws.conditional_formatting.add(score_range,
        ColorScaleRule(start_type='min', start_color=RED,
                       mid_type='percentile', mid_value=50, mid_color=YELLOW,
                       end_type='max', end_color=GREEN))

    set_widths(ws, [30, 10, 12, 12, 12, 30])


# ============================================================
# GABARIT 13 : REGISTRE ANOMALIES BCF
# ============================================================
def gabarit_anomalies(wb):
    ws = wb.create_sheet("13. Anomalies BCF")

    style_titre(ws, 'A1:H1', "REGISTRE DES ANOMALIES (BCF)",
                bg=RED, fg=WHITE, taille=15)

    # KPI
    style_sous_titre(ws, 3, 'A3:H3', "SYNTHESE", bg=ORANGE)
    kpi_card(ws, 5, 1, "Total", "=COUNTA(A9:A50)-1", "", NAVY)
    kpi_card(ws, 5, 2, "Critiques", '=COUNTIF(D9:D50,"Critique")', "", RED)
    kpi_card(ws, 5, 3, "Majeures", '=COUNTIF(D9:D50,"Majeure")', "", ORANGE)
    kpi_card(ws, 5, 4, "Mineures", '=COUNTIF(D9:D50,"Mineure")', "", YELLOW)
    kpi_card(ws, 5, 5, "Resolues", '=COUNTIF(G9:G50,"Resolu")', "", GREEN)
    kpi_card(ws, 5, 6, "En cours", '=COUNTIF(G9:G50,"En cours")', "", NAVY)
    kpi_card(ws, 5, 7, "Ouvertes", '=COUNTIF(G9:G50,"Ouvert")', "", RED)
    kpi_card(ws, 5, 8, "% resolues",
             '=IFERROR(COUNTIF(G9:G50,"Resolu")/(COUNTA(G9:G50)-1),0)', "", GREEN)

    style_header(ws, 8, ["N°", "Date", "Titre anomalie", "Gravite",
                          "Localisation", "Attribue a", "Statut", "Commentaire"])

    from openpyxl.worksheet.datavalidation import DataValidation
    dv_gravite = DataValidation(type="list", formula1='"Critique,Majeure,Mineure"',
                                 allow_blank=True)
    dv_gravite.add("D9:D50")
    ws.add_data_validation(dv_gravite)

    dv_statut = DataValidation(type="list", formula1='"Ouvert,En cours,Resolu"',
                                allow_blank=True)
    dv_statut.add("G9:G50")
    ws.add_data_validation(dv_statut)

    # Exemples de lignes vides
    for r in range(9, 30):
        ws.cell(row=r, column=1, value=r - 8).alignment = Alignment(horizontal='center')
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = thin_border()

    # Format conditionnel gravite
    ws.conditional_formatting.add("D9:D50",
        CellIsRule(operator='equal', formula=['"Critique"'],
                   fill=PatternFill("solid", fgColor=RED),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add("D9:D50",
        CellIsRule(operator='equal', formula=['"Majeure"'],
                   fill=PatternFill("solid", fgColor=ORANGE),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add("D9:D50",
        CellIsRule(operator='equal', formula=['"Mineure"'],
                   fill=PatternFill("solid", fgColor=YELLOW),
                   font=Font(bold=True)))

    # Format conditionnel statut
    ws.conditional_formatting.add("G9:G50",
        CellIsRule(operator='equal', formula=['"Resolu"'],
                   fill=PatternFill("solid", fgColor="DCFCE7"),
                   font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add("G9:G50",
        CellIsRule(operator='equal', formula=['"Ouvert"'],
                   fill=PatternFill("solid", fgColor="FEE2E2"),
                   font=Font(color=RED, bold=True)))
    ws.conditional_formatting.add("G9:G50",
        CellIsRule(operator='equal', formula=['"En cours"'],
                   fill=PatternFill("solid", fgColor="FEF3C7")))

    set_widths(ws, [6, 12, 35, 12, 20, 18, 12, 35])


# ============================================================
# GABARIT 14 : COURBE S - AVANCEMENT FINANCIER
# ============================================================
def gabarit_courbe_s(wb):
    ws = wb.create_sheet("14. Courbe S")

    style_titre(ws, 'A1:F1', "COURBE D'AVANCEMENT FINANCIER (S-CURVE)",
                bg=NAVY, taille=15)

    ws['A2'] = "Duree :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = 12
    ws['C2'] = "mois"

    style_header(ws, 4, ["Mois", "Prevu mensuel (€)", "Prevu cumule (€)",
                          "Realise mensuel (€)", "Realise cumule (€)", "Ecart (€)"])

    for i in range(1, 13):
        r = 4 + i
        ws.cell(row=r, column=1, value="M" + str(i)).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=0).number_format = '#,##0 €'
        if i == 1:
            ws.cell(row=r, column=3, value="=B" + str(r))
        else:
            ws.cell(row=r, column=3, value="=C" + str(r - 1) + "+B" + str(r))
        ws.cell(row=r, column=3).number_format = '#,##0 €'
        ws.cell(row=r, column=4, value=0).number_format = '#,##0 €'
        if i == 1:
            ws.cell(row=r, column=5, value="=D" + str(r))
        else:
            ws.cell(row=r, column=5, value="=E" + str(r - 1) + "+D" + str(r))
        ws.cell(row=r, column=5).number_format = '#,##0 €'
        ws.cell(row=r, column=6, value="=E" + str(r) + "-C" + str(r))
        ws.cell(row=r, column=6).number_format = '#,##0 €;[Red]-#,##0 €'
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border()

    # Graphique courbe S
    chart = LineChart()
    chart.title = "Courbe S : Prevu vs Realise"
    chart.y_axis.title = 'Cumul € HT'
    chart.x_axis.title = 'Mois'
    data = Reference(ws, min_col=3, min_row=4, max_col=3, max_row=16)
    data2 = Reference(ws, min_col=5, min_row=4, max_col=5, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data2, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=5, max_row=16)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws.add_chart(chart, "H4")

    # Format conditionnel ecart
    ws.conditional_formatting.add("F5:F16",
        ColorScaleRule(start_type='min', start_color=RED,
                       mid_type='num', mid_value=0, mid_color=WHITE,
                       end_type='max', end_color=GREEN))

    set_widths(ws, [8, 18, 18, 18, 18, 18])


# ============================================================
# GABARIT 15 : CONVENTION BIM - LOD / RACI
# ============================================================
def gabarit_convention_bim(wb):
    ws = wb.create_sheet("15. Convention BIM")

    style_titre(ws, 'A1:F1', "CONVENTION BIM - LOD & RACI",
                bg=TURQUOISE, fg=WHITE, taille=15)

    ws['A2'] = "Projet :"
    ws['A2'].font = Font(bold=True, color=NAVY)
    ws['B2'] = "[A remplir]"

    # LOD
    style_sous_titre(ws, 4, 'A4:F4',
                     "NIVEAUX DE DEVELOPPEMENT (LOD)", bg=ORANGE)
    style_header(ws, 5, ["Element", "Phase ESQ/APS", "Phase APD",
                           "Phase PRO", "Phase EXE", "Phase DOE"])

    elements = [
        "Structure (murs, dalles, poteaux)",
        "Menuiseries ext. (portes, fenetres)",
        "Cloisons / Doublages",
        "Reseaux plomberie",
        "Reseaux electriques",
        "CVC (chauffage, ventilation)",
        "Revetements",
        "Mobilier / equipements",
        "Terrain / topographie",
        "VRD exterieurs",
    ]
    lod_defaut = [
        [200, 300, 350, 400, 500],
        [200, 300, 350, 400, 500],
        [100, 200, 300, 350, 400],
        [100, 200, 300, 400, 500],
        [100, 200, 300, 400, 500],
        [100, 200, 300, 400, 500],
        [100, 200, 300, 350, 400],
        [100, 100, 200, 300, 400],
        [200, 200, 300, 300, 300],
        [100, 200, 300, 350, 400],
    ]

    for i, (elem, lods) in enumerate(zip(elements, lod_defaut)):
        r = 6 + i
        ws.cell(row=r, column=1, value=elem)
        for j, lod in enumerate(lods):
            c = ws.cell(row=r, column=2 + j, value="LOD " + str(lod))
            c.alignment = Alignment(horizontal='center')
            # Couleur selon LOD
            if lod <= 200:
                c.fill = PatternFill("solid", fgColor="DBEAFE")
            elif lod <= 300:
                c.fill = PatternFill("solid", fgColor="FEF3C7")
            elif lod <= 400:
                c.fill = PatternFill("solid", fgColor="FED7AA")
            else:
                c.fill = PatternFill("solid", fgColor="DCFCE7")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = thin_border()

    # RACI
    start_raci = 6 + len(elements) + 2
    style_sous_titre(ws, start_raci, 'A' + str(start_raci) + ':F' + str(start_raci),
                     "MATRICE RACI (R=Responsable A=Accountable C=Consulte I=Informe)",
                     bg=ORANGE)
    style_header(ws, start_raci + 1,
                 ["Tache BIM", "MOA", "MOE archi", "BET struct", "BIM Manager", "Entreprise"])

    taches_raci = [
        "Definition convention BIM",
        "Maquette architecturale",
        "Maquette structure",
        "Maquette fluides",
        "Controle qualite maquette",
        "Synthese / detection de clash",
        "Extraction quantites",
        "Maquette DOE",
        "Plan execution",
        "Suivi de chantier BIM",
    ]

    from openpyxl.worksheet.datavalidation import DataValidation
    dv_raci = DataValidation(type="list", formula1='"R,A,C,I,-"', allow_blank=True)

    for i, tache in enumerate(taches_raci):
        r = start_raci + 2 + i
        ws.cell(row=r, column=1, value=tache)
        for col in range(2, 7):
            ws.cell(row=r, column=col, value="").alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=col).border = thin_border()
        ws.cell(row=r, column=1).border = thin_border()

    dv_raci.add("B" + str(start_raci + 2) + ":F" + str(start_raci + 1 + len(taches_raci)))
    ws.add_data_validation(dv_raci)

    # Couleurs RACI
    raci_range = "B" + str(start_raci + 2) + ":F" + str(start_raci + 1 + len(taches_raci))
    ws.conditional_formatting.add(raci_range,
        CellIsRule(operator='equal', formula=['"R"'],
                   fill=PatternFill("solid", fgColor=RED),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(raci_range,
        CellIsRule(operator='equal', formula=['"A"'],
                   fill=PatternFill("solid", fgColor=ORANGE),
                   font=Font(color=WHITE, bold=True)))
    ws.conditional_formatting.add(raci_range,
        CellIsRule(operator='equal', formula=['"C"'],
                   fill=PatternFill("solid", fgColor=YELLOW),
                   font=Font(bold=True)))
    ws.conditional_formatting.add(raci_range,
        CellIsRule(operator='equal', formula=['"I"'],
                   fill=PatternFill("solid", fgColor=TURQUOISE),
                   font=Font(color=WHITE, bold=True)))

    set_widths(ws, [35, 12, 12, 12, 14, 14])


# ============================================================
# PAGE GARDE
# ============================================================
def page_garde(wb):
    ws = wb.active
    ws.title = "Sommaire"

    style_titre(ws, 'A1:D1', "GABARITS BIM - BTS MEC 2026",
                bg=NAVY, taille=22, hauteur=60)

    ws['A3'] = "Candidat :"
    ws['A3'].font = Font(bold=True, size=12, color=NAVY)
    ws['B3'] = "BAHAFID Mohamed"
    ws['A4'] = "Epreuve :"
    ws['A4'].font = Font(bold=True, size=12, color=NAVY)
    ws['B4'] = "E6-A Projet numerique BIM"
    ws['A5'] = "Date :"
    ws['A5'].font = Font(bold=True, size=12, color=NAVY)
    ws['B5'] = "[a remplir]"

    style_sous_titre(ws, 7, 'A7:D7', "SOMMAIRE DES GABARITS", bg=ORANGE)

    gabarits = [
        ("1", "Dashboard projet", "KPI globaux, repartition lots, graphique donut"),
        ("2", "DPGF professionnel", "12 lots, totaux HT/TVA/TTC automatiques"),
        ("3", "Planning chantier", "Gantt 12 mois avec remplissage auto"),
        ("4", "Bilan carbone RE2020", "IC construction, seuils, graphique"),
        ("5", "Controle qualite BIM", "Checklist 30 points, score conformite"),
        ("6", "Metre detaille", "Calculs L x l x h par ouvrage"),
        ("7", "DQE estimatif", "Devis quantitatif avec prix mat./pose"),
        ("8", "Sous-detail de prix", "Etude de prix : materiaux + MO + materiel"),
        ("9", "Note de synthese", "Rapport de projet en 10 sections"),
        ("10", "Descente de charges", "G + Q, combinaisons ELU / ELS"),
        ("11", "Bilan thermique", "Bbio, Cep, Ic - seuils RE2020"),
        ("12", "Comparatif variantes", "Analyse multicritere ponderee"),
        ("13", "Registre anomalies BCF", "Suivi des incompatibilites (gravite, statut)"),
        ("14", "Courbe S", "Avancement financier prevu vs realise"),
        ("15", "Convention BIM + RACI", "LOD par phase, matrice responsabilites"),
    ]

    style_header(ws, 9, ["N°", "Gabarit", "Contenu", ""])

    for i, (num, nom, contenu) in enumerate(gabarits):
        r = 10 + i
        ws.cell(row=r, column=1, value=num).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=1).font = Font(bold=True, size=14, color=ORANGE)
        ws.cell(row=r, column=2, value=nom).font = Font(bold=True, color=NAVY)
        ws.cell(row=r, column=3, value=contenu)
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = thin_border()
        ws.row_dimensions[r].height = 24

    r = 10 + len(gabarits) + 2
    ws.cell(row=r, column=1, value="Mode d'emploi :").font = Font(bold=True, size=12, color=NAVY)
    r += 1
    ws.cell(row=r, column=1, value="1. Completer les champs [A remplir] en en-tete de chaque onglet")
    r += 1
    ws.cell(row=r, column=1, value="2. Saisir les quantites extraites de la maquette eveBIM")
    r += 1
    ws.cell(row=r, column=1, value="3. Les totaux, pourcentages et indicateurs se calculent automatiquement")
    r += 1
    ws.cell(row=r, column=1, value="4. Les graphiques et jauges se mettent a jour en temps reel")
    r += 1
    ws.cell(row=r, column=1, value="5. Imprimer en A4 paysage pour les annexes")
    r += 2
    ws.cell(row=r, column=1, value="Astuce examen :").font = Font(bold=True, size=12, color=ORANGE)
    r += 1
    ws.cell(row=r, column=1, value="Utiliser les scripts 2_extract_quantites.py et 3_generer_dpgf.py")
    r += 1
    ws.cell(row=r, column=1, value="pour remplir automatiquement ces gabarits a partir du fichier IFC.")

    set_widths(ws, [8, 30, 55, 5])


# ============================================================
# MAIN
# ============================================================
def generer_gabarits(dossier_sortie):
    if not os.path.isdir(dossier_sortie):
        os.makedirs(dossier_sortie, exist_ok=True)

    wb = Workbook()
    page_garde(wb)
    gabarit_dashboard(wb)
    gabarit_dpgf(wb)
    gabarit_planning(wb)
    gabarit_bilan_carbone(wb)
    gabarit_controle_qualite(wb)
    gabarit_metre(wb)
    gabarit_dqe(wb)
    gabarit_sous_detail(wb)
    gabarit_note_synthese(wb)
    gabarit_descente_charges(wb)
    gabarit_bilan_thermique(wb)
    gabarit_variantes(wb)
    gabarit_anomalies(wb)
    gabarit_courbe_s(wb)
    gabarit_convention_bim(wb)

    output = os.path.join(dossier_sortie, "GABARITS_BIM_Examen.xlsx")

    try:
        wb.save(output)
    except Exception as e:
        print("[ERREUR] Impossible d'ecrire : " + str(e))
        sys.exit(1)

    print("")
    print("[OK] Gabarits generes : " + output)
    print("")
    print("Contenu (15 gabarits) :")
    print("  Sommaire")
    print("  1. Dashboard projet            9. Note de synthese")
    print("  2. DPGF pro                    10. Descente de charges")
    print("  3. Planning Gantt              11. Bilan thermique RE2020")
    print("  4. Bilan carbone RE2020        12. Comparatif variantes")
    print("  5. Controle qualite BIM        13. Registre anomalies BCF")
    print("  6. Metre detaille              14. Courbe S")
    print("  7. DQE estimatif               15. Convention BIM + RACI")
    print("  8. Sous-detail de prix")
    print("")
    print("Pret a utiliser pour l'examen E6-A BIM.")


if __name__ == "__main__":
    dossier = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
    generer_gabarits(dossier)
