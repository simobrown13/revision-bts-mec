#!/usr/bin/env python3
"""
3_generer_dpgf.py — Generer un cadre DPGF (Decomposition du Prix Global et Forfaitaire)
Usage : python 3_generer_dpgf.py chemin/vers/maquette.ifc

Produit : fichier Excel avec un cadre DPGF vierge adapte a la maquette :
  - Structure par lots (GO, charpente, couverture, cloisons...)
  - Codification UNTEC
  - Unites conformes (m, m², m³, U, Ft, Ens)
  - Quantites extraites automatiquement
  - Colonnes PU HT et Total HT vides (a completer par l'entreprise)
"""
import sys
import os


def generer_dpgf(chemin_ifc: str):
    try:
        import ifcopenshell
        import ifcopenshell.util.element
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"[ERREUR] Bibliotheque manquante : {e}")
        sys.exit(1)

    if not os.path.exists(chemin_ifc):
        print(f"[ERREUR] Fichier introuvable : {chemin_ifc}")
        sys.exit(1)

    print(f"\nGeneration du cadre DPGF : {os.path.basename(chemin_ifc)}")
    model = ifcopenshell.open(chemin_ifc)

    def get_qto(element, prop_name):
        psets = ifcopenshell.util.element.get_psets(element, qtos_only=True)
        for qto_name, props in psets.items():
            if prop_name in props:
                return props[prop_name]
        return 0

    # Calculer les quantites principales
    vol_fond = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                   for e in (model.by_type("IfcFooting") + model.by_type("IfcPile")))
    vol_mur = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                  for e in model.by_type("IfcWall"))
    surf_mur = sum((get_qto(e, 'NetSideArea') or get_qto(e, 'GrossSideArea') or 0)
                   for e in model.by_type("IfcWall"))
    vol_dalle = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                    for e in model.by_type("IfcSlab"))
    surf_dalle = sum((get_qto(e, 'NetArea') or get_qto(e, 'GrossArea') or 0)
                     for e in model.by_type("IfcSlab"))
    vol_pot = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                  for e in model.by_type("IfcColumn"))
    vol_pou = sum((get_qto(e, 'NetVolume') or get_qto(e, 'GrossVolume') or 0)
                  for e in model.by_type("IfcBeam"))
    nb_portes = len(model.by_type("IfcDoor"))
    nb_fenetres = len(model.by_type("IfcWindow"))
    surf_toit = sum((get_qto(e, 'NetArea') or get_qto(e, 'GrossArea') or 0)
                    for e in model.by_type("IfcRoof"))

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DPGF"

    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    lot_font = Font(bold=True, size=12, color="FFFFFF")
    lot_fill = PatternFill("solid", fgColor="475569")
    sous_lot_fill = PatternFill("solid", fgColor="E2E8F0")
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # En-tete du document
    ws.merge_cells('A1:F1')
    ws['A1'] = "DECOMPOSITION DU PRIX GLOBAL ET FORFAITAIRE (DPGF)"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"Projet : {os.path.splitext(os.path.basename(chemin_ifc))[0]}"
    ws['A3'] = "Phase : DCE"

    # En-tetes colonnes
    headers = ["Code", "Designation", "Unite", "Quantite", "PU HT (EUR)", "Total HT (EUR)"]
    row = 5
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    row += 1

    def ajouter_lot(nom, couleur="475569"):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=nom)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=couleur)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

    def ajouter_ligne(code, designation, unite, quantite, bold=False):
        nonlocal row
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=designation)
        ws.cell(row=row, column=3, value=unite)
        ws.cell(row=row, column=4, value=round(quantite, 2) if isinstance(quantite, (int, float)) else quantite)
        # PU et Total vides (a completer)
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}" if isinstance(quantite, (int, float)) else "")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border_thin
            if bold:
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = sous_lot_fill
        row += 1

    # ===== LOT 01 — TERRASSEMENT =====
    ajouter_lot("LOT 01 - TERRASSEMENT", "8B5CF6")
    ajouter_ligne("01.01", "Decapage de la terre vegetale (ep. 30 cm)", "m²", surf_dalle * 1.2)
    ajouter_ligne("01.02", "Fouilles en pleine masse", "m³", vol_fond * 1.5)
    ajouter_ligne("01.03", "Fouilles en rigoles / tranchees", "m³", vol_fond * 0.3)
    ajouter_ligne("01.04", "Remblai compacte", "m³", vol_fond * 0.8)
    ajouter_ligne("01.05", "Evacuation des deblais", "m³", vol_fond * 1.2)

    # ===== LOT 02 — GROS OEUVRE =====
    ajouter_lot("LOT 02 - GROS OEUVRE", "3B82F6")

    ajouter_ligne("02.01", "Fondations - Semelles filantes (BA)", "m³", vol_fond * 0.6, bold=True)
    ajouter_ligne("02.02", "Fondations - Semelles isolees (BA)", "m³", vol_fond * 0.3)
    ajouter_ligne("02.03", "Fondations - Longrines (BA)", "m³", vol_fond * 0.1)

    ajouter_ligne("02.10", "Murs enterres BA (ep. 20 cm)", "m³", vol_mur * 0.15, bold=True)
    ajouter_ligne("02.11", "Dallage sur terre plein (ep. 12 cm)", "m²", surf_dalle * 0.5)

    ajouter_ligne("02.20", "Voiles BA (ep. 18 cm)", "m³", vol_mur * 0.6, bold=True)
    ajouter_ligne("02.21", "Poteaux BA", "m³", vol_pot)
    ajouter_ligne("02.22", "Poutres BA", "m³", vol_pou)
    ajouter_ligne("02.23", "Dalles BA", "m³", vol_dalle, bold=True)

    ajouter_ligne("02.30", "Maconnerie - Agglos porteurs 20 cm", "m²", surf_mur * 0.3)
    ajouter_ligne("02.31", "Maconnerie - Cloisons briques / BA13", "m²", surf_mur * 0.2)

    # ===== LOT 03 — CHARPENTE / COUVERTURE =====
    if surf_toit > 0:
        ajouter_lot("LOT 03 - CHARPENTE / COUVERTURE", "F59E0B")
        ajouter_ligne("03.01", "Charpente bois / metallique", "m²", surf_toit, bold=True)
        ajouter_ligne("03.02", "Couverture (tuiles, bac acier, etancheite)", "m²", surf_toit)
        ajouter_ligne("03.03", "Isolation rampants (200 mm laine)", "m²", surf_toit)
        ajouter_ligne("03.04", "Zinguerie / gouttieres / descentes", "ml", (surf_toit ** 0.5) * 4)

    # ===== LOT 04 — MENUISERIES EXTERIEURES =====
    if nb_portes + nb_fenetres > 0:
        ajouter_lot("LOT 04 - MENUISERIES EXTERIEURES", "22C55E")
        ajouter_ligne("04.01", "Portes exterieures (fourniture + pose)", "U", max(1, nb_portes // 3))
        ajouter_ligne("04.02", "Fenetres / baies (fourniture + pose)", "U", nb_fenetres, bold=True)
        ajouter_ligne("04.03", "Occultations (volets, stores)", "U", nb_fenetres)

    # ===== LOT 05 — CLOISONS / DOUBLAGES =====
    ajouter_lot("LOT 05 - CLOISONS / DOUBLAGES", "8B5CF6")
    ajouter_ligne("05.01", "Cloisons placo (72/48)", "m²", surf_mur * 0.4, bold=True)
    ajouter_ligne("05.02", "Doublages thermiques (PSE + BA13)", "m²", surf_mur * 0.3)
    ajouter_ligne("05.03", "Faux plafonds BA13", "m²", surf_dalle * 0.8)

    # ===== LOT 06 — REVETEMENTS DE SOLS =====
    ajouter_lot("LOT 06 - REVETEMENTS DE SOLS", "14B8A6")
    ajouter_ligne("06.01", "Chape beton / ragreage", "m²", surf_dalle)
    ajouter_ligne("06.02", "Carrelage (pieces humides)", "m²", surf_dalle * 0.2)
    ajouter_ligne("06.03", "Parquet / stratifie (chambres, sejour)", "m²", surf_dalle * 0.5, bold=True)
    ajouter_ligne("06.04", "Plinthes", "ml", (surf_dalle ** 0.5) * 8)

    # ===== LOT 07 — MENUISERIES INTERIEURES =====
    ajouter_lot("LOT 07 - MENUISERIES INTERIEURES", "F59E0B")
    ajouter_ligne("07.01", "Portes interieures (fourniture + pose)", "U", max(1, nb_portes - nb_portes // 3))
    ajouter_ligne("07.02", "Placards", "ml", nb_portes * 0.8)

    # ===== LOT 08 — PEINTURE =====
    ajouter_lot("LOT 08 - PEINTURE / REVETEMENTS MURAUX", "EF4444")
    ajouter_ligne("08.01", "Peinture murs / plafonds (2 couches)", "m²", surf_mur * 1.5, bold=True)
    ajouter_ligne("08.02", "Peinture boiseries", "m²", nb_portes * 4)

    # ===== LOT 09 — PLOMBERIE =====
    ajouter_lot("LOT 09 - PLOMBERIE / SANITAIRES", "3B82F6")
    ajouter_ligne("09.01", "Alimentation eau froide / eau chaude", "Ft", 1)
    ajouter_ligne("09.02", "Evacuations EU / EV", "Ft", 1)
    ajouter_ligne("09.03", "Appareils sanitaires (WC, lavabos, douches)", "Ens", 1)
    ajouter_ligne("09.04", "Production ECS (ballon / PAC / chauffe-eau)", "U", 1)

    # ===== LOT 10 — ELECTRICITE =====
    ajouter_lot("LOT 10 - ELECTRICITE COURANTS FORTS", "F59E0B")
    ajouter_ligne("10.01", "Tableau electrique + disjoncteurs", "U", 1)
    ajouter_ligne("10.02", "Points lumineux + interrupteurs", "U", surf_dalle / 15)
    ajouter_ligne("10.03", "Prises de courant", "U", surf_dalle / 10)

    # ===== LOT 11 — CHAUFFAGE / VENTILATION =====
    ajouter_lot("LOT 11 - CHAUFFAGE / VENTILATION / CLIMATISATION", "EF4444")
    ajouter_ligne("11.01", "Systeme de chauffage (PAC, chaudiere, radiateurs)", "Ens", 1)
    ajouter_ligne("11.02", "Ventilation (VMC simple ou double flux)", "Ens", 1)
    ajouter_ligne("11.03", "Gaines et bouches", "Ft", 1)

    # ===== LOT 12 — VRD / AMENAGEMENTS EXTERIEURS =====
    ajouter_lot("LOT 12 - VRD / AMENAGEMENTS EXTERIEURS", "22C55E")
    ajouter_ligne("12.01", "Voiries / acces", "m²", surf_dalle * 0.3)
    ajouter_ligne("12.02", "Reseaux exterieurs (eaux, electricite, telecom)", "Ft", 1)
    ajouter_ligne("12.03", "Espaces verts / plantations", "m²", surf_dalle * 0.5)

    # ===== TOTAL =====
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="TOTAL HT")
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="991B1B")
    cell.alignment = Alignment(horizontal='right', vertical='center', indent=2)
    ws.row_dimensions[row].height = 28

    total_cell = ws.cell(row=row, column=6, value=f"=SUM(F6:F{row-1})")
    total_cell.font = Font(bold=True, size=14)
    total_cell.fill = PatternFill("solid", fgColor="FEF3C7")
    total_cell.number_format = '#,##0.00 EUR'
    row += 1

    # Largeurs colonnes
    for col, width in enumerate([10, 45, 10, 15, 15, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Sauvegarder
    base = os.path.splitext(chemin_ifc)[0]
    output = base + "_DPGF.xlsx"
    wb.save(output)

    print(f"\n[OK] DPGF genere : {output}")
    print(f"\nResume des quantites utilisees :")
    print(f"  Fondations   : {vol_fond:.2f} m³")
    print(f"  Murs         : {vol_mur:.2f} m³ ({surf_mur:.2f} m²)")
    print(f"  Dalles       : {vol_dalle:.2f} m³ ({surf_dalle:.2f} m²)")
    print(f"  Poteaux      : {vol_pot:.2f} m³")
    print(f"  Poutres      : {vol_pou:.2f} m³")
    print(f"  Portes       : {nb_portes}")
    print(f"  Fenetres     : {nb_fenetres}")
    print(f"  Toiture      : {surf_toit:.2f} m²")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python 3_generer_dpgf.py <chemin_vers_maquette.ifc>")
        sys.exit(1)
    generer_dpgf(sys.argv[1])
