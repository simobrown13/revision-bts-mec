#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5_comparer_ifc.py - Comparer deux versions d'une maquette IFC
Usage : python 5_comparer_ifc.py maquette_v1.ifc maquette_v2.ifc
"""
import sys
import os

from _utils import setup_encoding, check_dependencies, safe_volume, safe_area

setup_encoding()
check_dependencies()

import ifcopenshell
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


TYPES_COMPARES = [
    "IfcWall", "IfcSlab", "IfcBeam", "IfcColumn",
    "IfcFooting", "IfcDoor", "IfcWindow", "IfcRoof",
    "IfcStair", "IfcRailing", "IfcCovering",
]


def verifier_fichiers(chemin_v1, chemin_v2):
    """Verifie que les deux fichiers existent."""
    if not os.path.exists(chemin_v1):
        print("[ERREUR] Fichier V1 introuvable : " + chemin_v1)
        sys.exit(1)
    if not os.path.exists(chemin_v2):
        print("[ERREUR] Fichier V2 introuvable : " + chemin_v2)
        sys.exit(1)


def ecrire_entete(ws, chemin_v1, chemin_v2):
    """Ecrit l'entete du rapport."""
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1E3A5F")

    ws.merge_cells('A1:F1')
    ws['A1'] = "COMPARAISON DE MAQUETTES IFC"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = "V1 : " + os.path.basename(chemin_v1)
    ws['A3'] = "V2 : " + os.path.basename(chemin_v2)


def ecrire_headers(ws, row, headers):
    """Ecrit les en-tetes d'un tableau."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


def colorer_ligne(ws, row, ecart):
    """Colore une ligne selon l'ecart."""
    if ecart > 0:
        fill = PatternFill("solid", fgColor="DCFCE7")
    elif ecart < 0:
        fill = PatternFill("solid", fgColor="FEE2E2")
    else:
        return

    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = fill


def comparer_nombres_elements(ws, m1, m2, row_start):
    """Compare le nombre d'elements par type entre V1 et V2."""
    ecrire_headers(ws, row_start, ["Type d'element", "Nombre V1", "Nombre V2",
                                    "Ecart", "Variation %", "Observation"])
    row = row_start + 1

    print("")
    print("%-20s %6s %6s %8s  Observation" % ("Type", "V1", "V2", "Ecart"))
    print("-" * 70)

    for type_ifc in TYPES_COMPARES:
        try:
            nb1 = len(m1.by_type(type_ifc))
            nb2 = len(m2.by_type(type_ifc))
        except Exception:
            continue

        if nb1 == 0 and nb2 == 0:
            continue

        ecart = nb2 - nb1
        variation = (ecart / nb1 * 100) if nb1 > 0 else (100 if nb2 > 0 else 0)

        if ecart > 0:
            obs = "+%d ajoute(s)" % ecart
        elif ecart < 0:
            obs = "%d supprime(s)" % ecart
        else:
            obs = "Identique"

        ws.cell(row=row, column=1, value=type_ifc)
        ws.cell(row=row, column=2, value=nb1)
        ws.cell(row=row, column=3, value=nb2)
        ws.cell(row=row, column=4, value=ecart)
        ws.cell(row=row, column=5, value="%+.1f%%" % variation)
        ws.cell(row=row, column=6, value=obs)
        colorer_ligne(ws, row, ecart)

        print("%-20s %6d %6d %+8d  %s" % (type_ifc, nb1, nb2, ecart, obs))
        row += 1

    return row


def comparer_quantites(ws, m1, m2, row_start):
    """Compare les quantites (volumes et surfaces) entre V1 et V2."""
    # Titre
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1E3A5F")
    row = row_start + 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="COMPARAISON DES QUANTITES")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1

    ecrire_headers(ws, row, ["Grandeur", "Unite", "V1", "V2", "Ecart", "Variation %"])
    row += 1

    def total_vol(model, type_ifc):
        return sum(safe_volume(e) for e in model.by_type(type_ifc))

    def total_surf(model, type_ifc):
        return sum(safe_area(e) for e in model.by_type(type_ifc))

    grandeurs = [
        ("Volume murs", "m³", total_vol(m1, "IfcWall"), total_vol(m2, "IfcWall")),
        ("Volume dalles", "m³", total_vol(m1, "IfcSlab"), total_vol(m2, "IfcSlab")),
        ("Volume poteaux", "m³", total_vol(m1, "IfcColumn"), total_vol(m2, "IfcColumn")),
        ("Volume poutres", "m³", total_vol(m1, "IfcBeam"), total_vol(m2, "IfcBeam")),
        ("Volume fondations", "m³",
            total_vol(m1, "IfcFooting") + total_vol(m1, "IfcPile"),
            total_vol(m2, "IfcFooting") + total_vol(m2, "IfcPile")),
        ("Surface dalles", "m²", total_surf(m1, "IfcSlab"), total_surf(m2, "IfcSlab")),
        ("Surface murs", "m²", total_surf(m1, "IfcWall"), total_surf(m2, "IfcWall")),
    ]

    yellow = PatternFill("solid", fgColor="FEF3C7")

    for nom, unite, v1, v2 in grandeurs:
        if v1 == 0 and v2 == 0:
            continue

        ecart = v2 - v1
        variation = (ecart / v1 * 100) if v1 > 0 else 100

        ws.cell(row=row, column=1, value=nom)
        ws.cell(row=row, column=2, value=unite)
        ws.cell(row=row, column=3, value=round(v1, 2))
        ws.cell(row=row, column=4, value=round(v2, 2))
        ws.cell(row=row, column=5, value=round(ecart, 2))
        ws.cell(row=row, column=6, value="%+.1f%%" % variation)

        if abs(variation) > 10:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = yellow
        row += 1

    return row


def comparer_ifc(chemin_v1, chemin_v2):
    verifier_fichiers(chemin_v1, chemin_v2)

    print("")
    print("Comparaison :")
    print("  V1 : " + os.path.basename(chemin_v1))
    print("  V2 : " + os.path.basename(chemin_v2))

    try:
        m1 = ifcopenshell.open(chemin_v1)
        m2 = ifcopenshell.open(chemin_v2)
    except Exception as e:
        print("[ERREUR] " + str(e))
        sys.exit(1)

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison"

    ecrire_entete(ws, chemin_v1, chemin_v2)
    row = comparer_nombres_elements(ws, m1, m2, 5)
    comparer_quantites(ws, m1, m2, row)

    # Largeurs
    widths = [25, 12, 12, 12, 12, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # Sauvegarder
    base = os.path.splitext(chemin_v1)[0]
    output = base + "_vs_v2_comparaison.xlsx"

    try:
        wb.save(output)
    except Exception as e:
        print("[ERREUR] Impossible d'ecrire : " + str(e))
        sys.exit(1)

    print("")
    print("[OK] Comparaison generee : " + output)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage : python 5_comparer_ifc.py <v1.ifc> <v2.ifc>")
        sys.exit(1)
    comparer_ifc(sys.argv[1], sys.argv[2])
